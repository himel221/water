from django.shortcuts import render, get_object_or_404, redirect 
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .models import Product, DistrictDeliveryCharge, Order, OrderItem
import json
import io
from decimal import Decimal
from datetime import datetime, timedelta
from django.db import models
from django.contrib import messages
from .forms import CustomerForm
from django.contrib.auth import login as auth_login

from django.contrib.auth import get_user_model
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import Product, Inventory
import json
User = get_user_model()

# ReportLab Import
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️ ReportLab not installed. Please install: pip install reportlab")


# ===== Page Views =====

def home(request):
    """Home page view"""
    return render(request, 'home.html', {'active_page': 'home'})


def shop(request):
    """Shop page view - shows all active products"""
    products = Product.objects.filter(is_active=True)
    
    # Get unique categories for filter buttons
    categories = Product.objects.filter(is_active=True).values_list('category', flat=True).distinct()
    
    # Convert products to JSON for JavaScript
    products_json = []
    for product in products:
        original_price = float(product.price)
        effective_price = product.get_effective_price()
        discount_price = float(effective_price)
        discount_percent = product.discount_percentage or 0
        has_discount = product.is_on_sale and (product.discount_price is not None or discount_percent > 0)
        
        products_json.append({
            'id': product.id,
            'name': product.name,
            'description': product.description or '',
            'price': float(effective_price if has_discount else original_price),
            'discount_price': float(discount_price) if has_discount else None,
            'original_price': float(original_price),
            'discount_percentage': int(discount_percent) if has_discount else 0,
            'is_on_sale': bool(has_discount),
            'savings': float(original_price - discount_price) if has_discount else 0,
            'special_offer': product.special_offer or '',
            'category': product.category,
            'benefits': product.get_benefits_list(),
            'rating': float(product.rating) if product.rating else 4.5,
            'reviews': int(product.reviews) if product.reviews else 0,
            'stock_quantity': int(product.stock_quantity) if product.stock_quantity else 0,
            'is_in_stock': bool(product.stock_quantity > 0) if product.stock_quantity else False,
            'image': product.image.url if product.image else '/static/images/default-product.jpg',
            'is_active': bool(product.is_active),
        })
    
    return render(request, 'shop.html', {
        'active_page': 'shop',
        'products': products,
        'categories': categories,
        'products_json': json.dumps(products_json)
    })


def science(request):
    """Science page view"""
    return render(request, 'science.html', {'active_page': 'science'})


def explore(request):
    """Explore page view"""
    return render(request, 'explore.html', {'active_page': 'explore'})


def cart(request):
    """Cart page view"""
    return render(request, 'cart.html', {'active_page': 'cart'})


# ===== API Views =====

def product_list(request):
    """API: Get all active products as JSON"""
    products = Product.objects.filter(is_active=True)
    data = []
    for product in products:
        original_price = float(product.price)
        effective_price = product.get_effective_price()
        discount_price = float(effective_price)
        discount_percent = product.discount_percentage or 0
        has_discount = product.is_on_sale and (product.discount_price is not None or discount_percent > 0)
        
        data.append({
            'id': product.id,
            'name': product.name,
            'description': product.description or '',
            'price': float(effective_price if has_discount else original_price),
            'original_price': float(original_price),
            'discount_price': float(discount_price) if has_discount else None,
            'discount_percentage': int(discount_percent) if has_discount else 0,
            'is_on_sale': bool(has_discount),
            'special_offer': product.special_offer or '',
            'category': product.category,
            'benefits': product.get_benefits_list(),
            'rating': float(product.rating) if product.rating else 4.5,
            'reviews': int(product.reviews) if product.reviews else 0,
            'image': product.image.url if product.image else '/static/images/default.jpg',
            'is_active': bool(product.is_active),
        })
    return JsonResponse(data, safe=False)


def checkout(request):
    """Checkout page view"""
    return render(request, 'checkout.html')


def district_delivery_charges(request):
    """API: Get all district delivery charges"""
    charges = DistrictDeliveryCharge.objects.filter(is_active=True)
    data = []
    for charge in charges:
        data.append({
            'id': charge.id,
            'district': charge.district,
            'district_name': charge.district_name,
            'charge': float(charge.charge),
            'delivery_time': charge.delivery_time,
        })
    return JsonResponse(data, safe=False)


def get_delivery_charge(request, district):
    """API: Get delivery charge for a specific district"""
    try:
        charge = DistrictDeliveryCharge.objects.filter(
            district__iexact=district, 
            is_active=True
        ).first()
        
        if charge:
            return JsonResponse({
                'status': 'success',
                'district': charge.district,
                'district_name': charge.district_name,
                'charge': float(charge.charge),
                'delivery_time': charge.delivery_time,
            })
        else:
            return JsonResponse({
                'status': 'error',
                'message': f'Delivery charge not found for district: {district}',
                'charge': 0,
                'delivery_time': 'Not available'
            }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'charge': 0,
            'delivery_time': 'Not available'
        }, status=500)


from django.db import models
from django.core.validators import RegexValidator
from django.utils import timezone

from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth import get_user_model
import json


@csrf_exempt
@require_http_methods(["POST"])
def create_order(request):
    """API: Create a new order - Requires user to be logged in"""
    
    if not request.user.is_authenticated:
        return JsonResponse({
            'status': 'error',
            'message': 'Please sign in to place your order',
            'redirect_url': '/login/?next=' + request.META.get('HTTP_REFERER', '/'),
            'requires_login': True
        }, status=401)
    
    try:
        data = json.loads(request.body)
        
        # ✅ ইউজার রিলোড করুন (get_user_model() ব্যবহার করুন)
        User = get_user_model()
        user = User.objects.get(id=request.user.id)
        
        # ✅ ফ্রন্টএন্ড থেকে ডেটা নিন
        customer_name = data.get('customer_name', '')
        customer_phone = data.get('customer_phone', '')
        customer_address = data.get('customer_address', '')
        district = data.get('customer_district', '')
        
        print(f"👤 Order by: {user.email}")
        print(f"👤 User ID: {user.id}")
        print(f"👤 Customer Name: {customer_name}")
        print(f"📱 Phone: {customer_phone}")
        print(f"📍 Address: {customer_address}")
        print(f"🏙️ District: {district}")
        
        # ডেলিভারি চার্জ
        delivery_charge_obj = DistrictDeliveryCharge.objects.filter(
            district__iexact=district, 
            is_active=True
        ).first()
        delivery_charge = float(delivery_charge_obj.charge) if delivery_charge_obj else 0
        delivery_zone = delivery_charge_obj.district_name if delivery_charge_obj else 'Unknown District'
        
        items_data = data.get('items', [])
        subtotal = 0
        total_savings = 0
        total_discount = 0
        
        with transaction.atomic():
            # Validate stock
            for item in items_data:
                product_id = item.get('id')
                quantity = int(item.get('quantity', 1))
                
                if product_id:
                    try:
                        product = Product.objects.select_for_update().get(id=product_id)
                        if product.stock_quantity < quantity:
                            return JsonResponse({
                                'status': 'error',
                                'message': f'Insufficient stock for {product.name}. Available: {product.stock_quantity}'
                            }, status=400)
                    except Product.DoesNotExist:
                        return JsonResponse({
                            'status': 'error',
                            'message': f'Product not found: {item.get("name", "Unknown")}'
                        }, status=400)
            
            # Calculate and reduce stock
            for item in items_data:
                product_id = item.get('id')
                price = float(item.get('price', 0))
                original_price = float(item.get('original_price', price))
                quantity = int(item.get('quantity', 1))
                
                subtotal += original_price * quantity
                savings = (original_price - price) * quantity
                if savings > 0:
                    total_savings += savings
                    total_discount += savings
                
                if product_id:
                    try:
                        product = Product.objects.get(id=product_id)
                        product.stock_quantity -= quantity
                        product.save()
                        print(f"✅ Product stock reduced: {product.name} - New stock: {product.stock_quantity}")
                    except Product.DoesNotExist:
                        pass
            
            total_amount = subtotal - total_savings + delivery_charge
            
            # ✅ Order তৈরি করুন - user_id ব্যবহার করুন
            order = Order.objects.create(
                customer_name=customer_name,
                customer_email=user.email,
                customer_phone=customer_phone,
                customer_address=customer_address,
                customer_district=district,
                delivery_charge=delivery_charge,
                delivery_zone=delivery_zone,
                subtotal=subtotal,
                total_savings=total_savings,
                total_discount=total_discount,
                total_amount=total_amount,
                payment_method=data.get('payment_method', 'Cash on Delivery'),
                transaction_id=data.get('transaction_id', ''),
                status='pending',
                payment_status='pending',
                user_id=user.id  # ✅ user.id ব্যবহার করুন
            )
            
            print(f"✅ Order created: {order.order_id}")
            
            # Create order items
            for item in items_data:
                price = float(item.get('price', 0))
                original_price = float(item.get('original_price', price))
                quantity = int(item.get('quantity', 1))
                product_id = item.get('id')
                
                product = None
                if product_id:
                    try:
                        product = Product.objects.get(id=product_id)
                    except Product.DoesNotExist:
                        pass
                
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    product_name=item.get('name', 'Unknown'),
                    product_price=price,
                    original_price=original_price,
                    quantity=quantity,
                    discount_percentage=int(item.get('discount_percentage', 0)),
                    special_offer=item.get('special_offer', ''),
                    total_price=price * quantity
                )
        
        return JsonResponse({
            'status': 'success',
            'order_id': order.order_id,
            'message': 'Order created successfully.',
            'subtotal': subtotal,
            'total_savings': total_savings,
            'total_discount': total_discount,
            'total_amount': total_amount,
            'user': user.email,
            'customer_name': customer_name
        })
        
    except json.JSONDecodeError as e:
        print(f"❌ JSON Decode Error: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid JSON data'
        }, status=400)
        
    except Exception as e:
        print(f"❌ Order Creation Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


def order_list(request):
    """API: Get all orders with items"""
    orders = Order.objects.all()
    data = []
    for order in orders:
        items_data = []
        for item in order.order_items.all():
            items_data.append({
                'product_name': item.product_name,
                'quantity': item.quantity,
                'product_price': float(item.product_price),
                'original_price': float(item.original_price),
                'discount_percentage': item.discount_percentage,
                'special_offer': item.special_offer,
                'total_price': float(item.total_price)
            })
        
        data.append({
            'order_id': order.order_id,
            'customer_name': order.customer_name,
            'customer_email': order.customer_email,
            'customer_phone': order.customer_phone,
            'customer_address': order.customer_address,
            'customer_district': order.customer_district,
            'delivery_zone': order.delivery_zone,
            'delivery_charge': float(order.delivery_charge),
            'items': items_data,
            'subtotal': float(order.subtotal),
            'total_savings': float(order.total_savings),
            'total_discount': float(order.total_discount or 0),
            'total_amount': float(order.total_amount),
            'status': order.status,
            'payment_status': order.payment_status,
            'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    return JsonResponse(data, safe=False)


def download_invoice(request, order_id):
    """Download invoice as PDF"""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("⚠️ ReportLab not installed. Please install: pip install reportlab", status=500)
    
    order = get_object_or_404(Order, order_id=order_id)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#007bff'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    heading_style = ParagraphStyle(
        'HeadingStyle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=10
    )
    
    normal_style = styles['Normal']
    
    story = []
    
    story.append(Paragraph("Aquanimity SuperWater - Invoice", title_style))
    story.append(Paragraph(f"Order ID: {order.order_id}", normal_style))
    story.append(Paragraph(f"Date: {order.created_at.strftime('%B %d, %Y at %I:%M %p')}", normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    story.append(Paragraph("Customer Details", heading_style))
    story.append(Paragraph(f"Name: {order.customer_name}", normal_style))
    story.append(Paragraph(f"Email: {order.customer_email}", normal_style))
    if order.customer_phone:
        story.append(Paragraph(f"Phone: {order.customer_phone}", normal_style))
    story.append(Paragraph(f"Address: {order.customer_address}", normal_style))
    if order.customer_district:
        story.append(Paragraph(f"District: {order.customer_district}", normal_style))
    story.append(Spacer(1, 0.2*inch))
    
    story.append(Paragraph("Order Items", heading_style))
    
    table_data = [
        ['#', 'Product', 'Qty', 'Unit Price', 'Original Price', 'Discount', 'Total']
    ]
    
    for idx, item in enumerate(order.order_items.all(), 1):
        discount_text = f"{item.discount_percentage}%" if item.discount_percentage > 0 else '-'
        if item.special_offer:
            discount_text += f" 🎉 {item.special_offer}"
        
        table_data.append([
            str(idx),
            item.product_name,
            str(item.quantity),
            f"৳ {float(item.product_price):.2f}",
            f"৳ {float(item.original_price):.2f}",
            discount_text,
            f"৳ {float(item.total_price):.2f}"
        ])
    
    table = Table(table_data, colWidths=[0.3*inch, 2.2*inch, 0.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    story.append(Paragraph("Order Summary", heading_style))
    
    summary_data = [
        ['Subtotal', f"৳ {float(order.subtotal):.2f}"],
    ]
    if order.total_savings > 0:
        summary_data.append(['Total Savings', f"৳ {float(order.total_savings):.2f}"])
    if order.delivery_charge > 0:
        summary_data.append(['Delivery Charge', f"৳ {float(order.delivery_charge):.2f}"])
    summary_data.append(['Total Amount', f"৳ {float(order.total_amount):.2f}"])
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4fd')),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph("Thank you for shopping with Aquanimity SuperWater!", footer_style))
    story.append(Paragraph("© 2026 Aquanimity SuperWater. All rights reserved.", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{order.order_id}.pdf"'
    return response


# ===== Admin Views =====

def admin_dashboard(request):
    """Admin Dashboard View"""
    from .models import Order, Product, OrderItem, DistrictDeliveryCharge
    
    context = {
        'active_page': 'admin_dashboard',
        'total_orders': Order.objects.count(),
        'total_products': Product.objects.count(),
        'total_customers': Order.objects.values('customer_email').distinct().count(),
        'total_districts': DistrictDeliveryCharge.objects.filter(is_active=True).count(),
        'total_revenue': sum(float(o.total_amount) for o in Order.objects.all()),
        'recent_orders': Order.objects.all().order_by('-created_at')[:5],
        'pending_orders': Order.objects.filter(status='pending').count(),
        'shipped_orders': Order.objects.filter(status='shipped').count(),
        'delivered_orders': Order.objects.filter(status='delivered').count(),
        'total_items': OrderItem.objects.count(),
        'districts': DistrictDeliveryCharge.objects.filter(is_active=True),
        'products': Product.objects.filter(is_active=True)[:5],
    }
    return render(request, 'admin_dashboard.html', context)


def admin_orders(request):
    """Admin Orders Management View"""
    from .models import Order, OrderItem, Product
    import json
    
    orders = Order.objects.all().order_by('-created_at')
    
    total_orders = orders.count()
    pending_orders = orders.filter(status='pending').count()
    shipped_orders = orders.filter(status='shipped').count()
    delivered_orders = orders.filter(status='delivered').count()
    cancelled_orders = orders.filter(status='cancelled').count()
    
    orders_json = []
    for order in orders:
        total_discount = float(order.total_discount) if order.total_discount else 0
        
        special_offers = []
        for item in order.order_items.all():
            if item.special_offer:
                special_offers.append(item.special_offer)
        special_offers = list(set(special_offers))
        
        orders_json.append({
            'order_id': order.order_id,
            'customer_name': order.customer_name,
            'customer_email': order.customer_email,
            'customer_phone': order.customer_phone or '',
            'customer_address': order.customer_address,
            'customer_district': order.customer_district or '',
            'get_total_items': order.get_total_items(),
            'total_amount': float(order.total_amount) if order.total_amount else 0,
            'total_savings': float(order.total_savings) if order.total_savings else 0,
            'total_discount': total_discount,
            'status': order.status,
            'special_offers_list': special_offers,
            'is_eligible_for_offer': order.is_eligible_for_offer(),
            'created_at': order.created_at.isoformat(),
            'items': [
                {
                    'product_name': item.product_name,
                    'quantity': item.quantity,
                    'product_price': float(item.product_price),
                    'original_price': float(item.original_price),
                    'discount_percentage': item.discount_percentage,
                    'special_offer': item.special_offer or '',
                }
                for item in order.order_items.all()
            ]
        })
    
    context = {
        'active_page': 'admin_orders',
        'orders': orders,
        'orders_json': json.dumps(orders_json) if orders_json else '[]',
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'total_products': Product.objects.count(),
        'total_customers': orders.values('customer_email').distinct().count(),
    }
    return render(request, 'admin_orders.html', context)


# ===== Export Functions =====

import io
import openpyxl
from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Q
from django.contrib.admin.views.decorators import staff_member_required
import json

from water_app.models import Order, OrderItem, Product  # Removed SpecialOffer

# Try to import ReportLab
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4, portrait
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


@staff_member_required
def export_orders_excel(request):
    """Export orders as Excel file with ALL fields"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("⚠️ openpyxl not installed. Please install: pip install openpyxl", status=500)
    
    # Get orders based on filters
    orders = Order.objects.all().order_by('-created_at')
    
    # Apply filters if provided
    status = request.GET.get('status')
    date = request.GET.get('date')
    ids = request.GET.get('ids')
    
    if ids:
        order_ids = ids.split(',')
        orders = orders.filter(order_id__in=order_ids)
    elif status and status != 'all':
        orders = orders.filter(status=status)
    
    # Date filter
    if date and date != 'all':
        now = datetime.now().date()
        if date == 'today':
            orders = orders.filter(created_at__date=now)
        elif date == 'week':
            week_ago = now - timedelta(days=7)
            orders = orders.filter(created_at__date__gte=week_ago)
        elif date == 'month':
            orders = orders.filter(created_at__year=now.year, created_at__month=now.month)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Report"
    
    # Styles
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sub_header_fill = PatternFill(start_color='2e86c1', end_color='2e86c1', fill_type='solid')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # ===== Title Section =====
    ws.merge_cells('A1:S1')
    title_cell = ws['A1']
    title_cell.value = "📦 Aquanimity SuperWater - Complete Orders Report"
    title_cell.font = Font(name='Arial', size=18, bold=True, color='1a5276')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:S2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total Orders: {orders.count()}"
    subtitle_cell.font = Font(name='Arial', size=10, color='666666')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # ===== Summary Statistics =====
    summary_row = 4
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    summary_title = ws.cell(row=summary_row, column=1)
    summary_title.value = "📊 Summary Statistics"
    summary_title.font = Font(name='Arial', size=12, bold=True, color='1a5276')
    
    total_revenue = sum(float(o.total_amount) for o in orders)
    total_discount = sum(float(o.total_discount or 0) for o in orders)
    total_savings = sum(float(o.total_savings or 0) for o in orders)
    
    # Check if orders have is_eligible_for_offer method
    eligible_count = 0
    for o in orders:
        try:
            if o.is_eligible_for_offer():
                eligible_count += 1
        except:
            pass
    
    summary_data = [
        ['Total Orders', str(orders.count())],
        ['Total Revenue', f"৳ {total_revenue:.2f}"],
        ['Total Discount Given', f"৳ {total_discount:.2f}"],
        ['Total Savings', f"৳ {total_savings:.2f}"],
        ['Eligible for Offer', str(eligible_count)],
        ['Not Eligible', str(orders.count() - eligible_count)],
    ]
    
    for idx, (label, value) in enumerate(summary_data):
        row = summary_row + idx + 1
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill(start_color='eaf2f8', end_color='eaf2f8', fill_type='solid')
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.border = border
        value_cell.alignment = Alignment(horizontal='left', vertical='center')
        if '৳' in value:
            value_cell.font = Font(color='007bff', bold=True)
    
    # ===== Main Headers - ALL FIELDS =====
    header_row = summary_row + len(summary_data) + 3
    
    # Complete headers list with ALL fields
    headers = [
        'Sl No', 
        'Order ID', 
        'Customer Name', 
        'Customer Email', 
        'Customer Phone', 
        'Customer Address', 
        'District', 
        'Total Items', 
        'Products Details',
        'Subtotal', 
        'Total Savings', 
        'Total Discount', 
        'Total Amount',
        'Status', 
        'Payment Method',
        'Payment Status',
        'Delivery Address',
        'Special Offers Applied',
        'Offer Eligibility',
        'Created Date',
        'Created Time'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # ===== Data Rows =====
    row_num = header_row + 1
    sl_no = 1
    
    for order in orders:
        # Get products details
        products = []
        special_offers = []
        
        for item in order.order_items.all():
            product_name = item.product_name or (item.product.name if item.product else 'Unknown')
            qty = item.quantity
            price = float(item.product_price or 0)
            original_price = float(item.original_price or price)
            discount_percent = float(item.discount_percentage or 0)
            
            product_detail = f"{product_name} x{qty}"
            if discount_percent > 0:
                product_detail += f" ({discount_percent}% OFF)"
            products.append(product_detail)
            
            # Special offers - check if field exists
            if hasattr(item, 'special_offer') and item.special_offer:
                special_offers.append(f"{product_name}: {item.special_offer}")
        
        products_text = '\n'.join(products[:10]) + ('\n...' if len(products) > 10 else '')
        special_offers_text = '\n'.join(special_offers) if special_offers else 'None'
        
        # Check eligibility - handle if method doesn't exist
        try:
            is_eligible = order.is_eligible_for_offer()
            eligibility = '✓ Eligible' if is_eligible else '✗ Not Eligible'
        except:
            eligibility = 'N/A'
        
        # Format date and time
        created_datetime = order.created_at
        created_date = created_datetime.strftime('%d %b %Y') if created_datetime else ''
        created_time = created_datetime.strftime('%I:%M %p') if created_datetime else ''
        
        # Payment status - check if field exists
        payment_status = 'N/A'
        if hasattr(order, 'payment_status') and order.payment_status:
            payment_status = order.payment_status
        
        row_data = [
            sl_no,
            order.order_id,
            order.customer_name,
            order.customer_email,
            order.customer_phone or '',
            order.customer_address or '',
            order.customer_district or '',
            order.get_total_items(),
            products_text,
            f"৳{float(order.subtotal or 0):.2f}",
            f"৳{float(order.total_savings or 0):.2f}",
            f"৳{float(order.total_discount or 0):.2f}",
            f"৳{float(order.total_amount):.2f}",
            order.get_status_display(),
            order.payment_method or 'Cash on Delivery',
            payment_status,
            getattr(order, 'delivery_address', '') or order.customer_address or '',
            special_offers_text,
            eligibility,
            created_date,
            created_time
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(
                horizontal='left' if col not in [1, 8, 10, 11, 12, 13] else 'center',
                vertical='center',
                wrap_text=True
            )
            cell.border = border
            
            # Color coding for status
            if col == 14:  # Status column
                status_colors = {
                    'pending': 'ffc107',
                    'processing': '17a2b8',
                    'shipped': '007bff',
                    'delivered': '28a745',
                    'cancelled': 'dc3545'
                }
                status_value = order.status
                if status_value in status_colors:
                    cell.fill = PatternFill(
                        start_color=status_colors[status_value],
                        end_color=status_colors[status_value],
                        fill_type='solid'
                    )
                    cell.font = Font(color='FFFFFF', bold=True)
            
            # Color coding for eligibility
            if col == 19:  # Eligibility column
                if 'Eligible' in value:
                    cell.fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                    cell.font = Font(color='155724', bold=True)
                elif 'Not' in value:
                    cell.fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                    cell.font = Font(color='721c24', bold=True)
            
            # Amount formatting
            if col in [10, 11, 12, 13]:  # Amount columns
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if col == 13:  # Total Amount - bold
                    cell.font = Font(bold=True, color='007bff')
                if col == 12 and float(order.total_discount or 0) > 0:  # Discount - red
                    cell.font = Font(color='dc3545')
        
        row_num += 1
        sl_no += 1
    
    # ===== Adjust Column Widths =====
    column_widths = {
        'A': 5,   # Sl No
        'B': 18,  # Order ID
        'C': 20,  # Customer Name
        'D': 28,  # Customer Email
        'E': 16,  # Customer Phone
        'F': 35,  # Customer Address
        'G': 14,  # District
        'H': 11,  # Total Items
        'I': 45,  # Products Details
        'J': 14,  # Subtotal
        'K': 16,  # Total Savings
        'L': 16,  # Total Discount
        'M': 16,  # Total Amount
        'N': 14,  # Status
        'O': 18,  # Payment Method
        'P': 16,  # Payment Status
        'Q': 35,  # Delivery Address
        'R': 30,  # Special Offers
        'S': 14,  # Eligibility
        'T': 14,  # Created Date
        'U': 12,  # Created Time
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ===== Freeze Panes =====
    ws.freeze_panes = f'A{header_row + 1}'
    
    # ===== Footer =====
    footer_row = row_num + 2
    ws.merge_cells(f'A{footer_row}:U{footer_row}')
    footer_cell = ws.cell(row=footer_row, column=1)
    footer_cell.value = "© 2026 Aquanimity Super Water. All rights reserved. | Report generated from Admin Dashboard"
    footer_cell.font = Font(name='Arial', size=9, color='999999')
    footer_cell.alignment = Alignment(horizontal='center')
    
    # ===== Response =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'orders_complete_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@staff_member_required
def export_orders_pdf(request):
    """Export orders as PDF file with ALL fields"""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("⚠️ ReportLab not installed. Please install: pip install reportlab", status=500)
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    # Get orders based on filters
    orders = Order.objects.all().order_by('-created_at')
    
    # Apply filters if provided
    status = request.GET.get('status')
    date = request.GET.get('date')
    ids = request.GET.get('ids')
    
    if ids:
        order_ids = ids.split(',')
        orders = orders.filter(order_id__in=order_ids)
    elif status and status != 'all':
        orders = orders.filter(status=status)
    
    # Date filter
    if date and date != 'all':
        now = datetime.now().date()
        if date == 'today':
            orders = orders.filter(created_at__date=now)
        elif date == 'week':
            week_ago = now - timedelta(days=7)
            orders = orders.filter(created_at__date__gte=week_ago)
        elif date == 'month':
            orders = orders.filter(created_at__year=now.year, created_at__month=now.month)
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f'orders_complete_report_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a5276'),
        alignment=TA_CENTER,
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        alignment=TA_CENTER,
        spaceAfter=16
    )
    
    heading_style = ParagraphStyle(
        'HeadingStyle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#1a5276'),
        alignment=TA_LEFT,
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9
    )
    
    header_cell_style = ParagraphStyle(
        'HeaderCellStyle',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        textColor=colors.white,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER
    )
    
    # Build story
    story = []
    
    # ===== Title =====
    story.append(Paragraph("📦 Aquanimity SuperWater - Complete Orders Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    story.append(Paragraph(f"Total Orders: {orders.count()} | Filter: {status if status != 'all' else 'All'} | Date: {date if date != 'all' else 'All Time'}", subtitle_style))
    story.append(Spacer(1, 0.2*inch))
    
    # ===== Summary Statistics =====
    total_revenue = sum(float(o.total_amount) for o in orders)
    total_discount = sum(float(o.total_discount or 0) for o in orders)
    total_savings = sum(float(o.total_savings or 0) for o in orders)
    
    # Check eligibility safely
    eligible_count = 0
    for o in orders:
        try:
            if o.is_eligible_for_offer():
                eligible_count += 1
        except:
            pass
    
    summary_data = [
        ['📊 SUMMARY STATISTICS', ''],
        ['Total Orders', str(orders.count())],
        ['Total Revenue', f"৳ {total_revenue:,.2f}"],
        ['Total Discount Given', f"৳ {total_discount:,.2f}"],
        ['Total Savings', f"৳ {total_savings:,.2f}"],
        ['Eligible for Offer', str(eligible_count)],
        ['Not Eligible', str(orders.count() - eligible_count)],
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 1), (-1, -1), 0.5, colors.HexColor('#ddd')),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#eaf2f8')),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#eaf2f8')),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # ===== Main Table Headers - ALL FIELDS =====
    # Complete headers with ALL fields
    headers = [
        'Sl', 'Order ID', 'Customer', 'Email', 'Phone', 'Address', 'District', 
        'Items', 'Products', 'Subtotal', 'Savings', 'Discount', 'Total', 
        'Status', 'Payment', 'Offers', 'Eligibility', 'Date'
    ]
    
    # Prepare table data
    table_data = [headers]
    
    # Track if we need pagination
    max_rows_per_page = 20
    total_orders = orders.count()
    display_orders = orders[:100] if total_orders > 100 else orders
    
    for idx, order in enumerate(display_orders, 1):
        # Get products
        products = []
        for item in order.order_items.all()[:5]:
            product_name = item.product_name or (item.product.name if item.product else 'Unknown')
            qty = item.quantity
            discount = f"({item.discount_percentage}%)" if item.discount_percentage > 0 else ''
            products.append(f"{product_name} x{qty}{discount}")
        products_text = ', '.join(products) + ('...' if order.order_items.count() > 5 else '')
        
        # Get offers - safely check if field exists
        offers = []
        for item in order.order_items.all():
            if hasattr(item, 'special_offer') and item.special_offer:
                offers.append(item.special_offer)
        offers_text = ', '.join(offers[:2]) + ('...' if len(offers) > 2 else '') if offers else '-'
        
        # Check eligibility safely
        try:
            is_eligible = order.is_eligible_for_offer()
            eligibility = '✓ Eligible' if is_eligible else '✗ Not'
        except:
            eligibility = 'N/A'
        
        # Format date
        created_date = order.created_at.strftime('%d/%m/%Y') if order.created_at else ''
        
        # Truncate long text
        address = (order.customer_address or '')[:25] + ('...' if len(order.customer_address or '') > 25 else '')
        
        table_data.append([
            str(idx),
            order.order_id,
            order.customer_name[:18] + ('...' if len(order.customer_name) > 18 else ''),
            order.customer_email[:20] + ('...' if len(order.customer_email) > 20 else ''),
            order.customer_phone or '-',
            address or '-',
            order.customer_district or '-',
            str(order.get_total_items()),
            products_text[:35] + ('...' if len(products_text) > 35 else ''),
            f"৳{float(order.subtotal or 0):.2f}",
            f"৳{float(order.total_savings or 0):.2f}",
            f"৳{float(order.total_discount or 0):.2f}",
            f"৳{float(order.total_amount):.2f}",
            order.get_status_display(),
            order.payment_method or 'COD',
            offers_text,
            eligibility,
            created_date
        ])
    
    if total_orders > 100:
        table_data.append([
            '...', f'... and {total_orders - 100} more orders', '', '', '', '', '', 
            '', '', '', '', '', '', '', '', '', '', ''
        ])
    
    # Calculate column widths
    col_widths = [
        0.7*inch, 0.5*inch, 0.7*inch, 1.0*inch, 0.7*inch, 0.9*inch, 0.4*inch,
        0.3*inch, .9*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch,
        0.7*inch, 0.6*inch, 0.5*inch, 0.6*inch, 0.6*inch
    ]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Apply table styles
    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        ('TOPPADDING', (0, 0), (-1, 0), 4),
        
        # Data rows
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 3),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ccc')),
        
        # Alternating row colors
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('BACKGROUND', (0, 2), (-1, -1), colors.HexColor('#ffffff')),
    ])
    
    # Add color coding for status column (column 13)
    for row_idx in range(1, len(table_data)):
        if len(table_data[row_idx]) > 13:
            status = table_data[row_idx][13]
            status_colors = {
                'Pending': '#ffc107',
                'Processing': '#17a2b8',
                'Shipped': '#007bff',
                'Delivered': '#28a745',
                'Cancelled': '#dc3545'
            }
            if status in status_colors:
                table_style.add('BACKGROUND', (13, row_idx), (13, row_idx), colors.HexColor(status_colors[status]))
                table_style.add('TEXTCOLOR', (13, row_idx), (13, row_idx), colors.white)
    
    # Color coding for eligibility (column 16)
    for row_idx in range(1, len(table_data)):
        if len(table_data[row_idx]) > 16:
            eligibility = table_data[row_idx][16]
            if 'Eligible' in eligibility:
                table_style.add('BACKGROUND', (16, row_idx), (16, row_idx), colors.HexColor('#d4edda'))
                table_style.add('TEXTCOLOR', (16, row_idx), (16, row_idx), colors.HexColor('#155724'))
            elif 'Not' in eligibility:
                table_style.add('BACKGROUND', (16, row_idx), (16, row_idx), colors.HexColor('#f8d7da'))
                table_style.add('TEXTCOLOR', (16, row_idx), (16, row_idx), colors.HexColor('#721c24'))
    
    table.setStyle(table_style)
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # ===== Footer =====
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER
    )
    story.append(Paragraph("© 2026 Aquanimity Super Water. All rights reserved. | Generated from Admin Dashboard", footer_style))
    story.append(Paragraph(f"Page 1 of 1 | Report includes {min(total_orders, 100)} of {total_orders} total orders", footer_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response


@staff_member_required
def download_invoice(request, order_id):
    """Download invoice as PDF for a specific order"""
    try:
        order = get_object_or_404(Order, order_id=order_id)
        # Assuming you have a function to generate invoice PDF
        from water_app.utils import generate_invoice_pdf
        return generate_invoice_pdf(order)
    except ImportError:
        # Fallback if utils doesn't exist
        return HttpResponse("Invoice generation not available", status=501)
    except Exception as e:
        return HttpResponse(f"Error generating invoice: {str(e)}", status=500)


@staff_member_required
@require_http_methods(["DELETE"])
def delete_order(request, order_id):
    """Delete an order"""
    try:
        order = get_object_or_404(Order, order_id=order_id)
        order.delete()
        return JsonResponse({'status': 'success', 'message': 'Order deleted successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@staff_member_required
def get_order_detail(request, order_id):
    """Get order details as JSON"""
    try:
        order = get_object_or_404(Order, order_id=order_id)
        data = {
            'order_id': order.order_id,
            'customer_name': order.customer_name,
            'customer_email': order.customer_email,
            'customer_phone': order.customer_phone,
            'customer_address': order.customer_address,
            'customer_district': order.customer_district,
            'total_amount': float(order.total_amount),
            'total_discount': float(order.total_discount or 0),
            'total_savings': float(order.total_savings or 0),
            'subtotal': float(order.subtotal or 0),
            'status': order.status,
            'status_display': order.get_status_display(),
            'payment_method': order.payment_method,
            'created_at': order.created_at.isoformat(),
            'is_eligible_for_offer': order.is_eligible_for_offer() if hasattr(order, 'is_eligible_for_offer') else False,
            'items': [
                {
                    'product_name': item.product_name or (item.product.name if item.product else 'Unknown'),
                    'quantity': item.quantity,
                    'product_price': float(item.product_price or 0),
                    'original_price': float(item.original_price or item.product_price or 0),
                    'discount_percentage': float(item.discount_percentage or 0),
                    'special_offer': getattr(item, 'special_offer', None),
                }
                for item in order.order_items.all()
            ],
            'special_offers_list': [
                getattr(item, 'special_offer', None) for item in order.order_items.all() 
                if getattr(item, 'special_offer', None)
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    

import json
import io
from datetime import datetime
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.core.files.storage import default_storage
from decimal import Decimal

from water_app.models import Order, OrderItem, Product


# ============================================
# PRODUCT MANAGEMENT VIEWS
# ============================================
@staff_member_required
def admin_products(request):
    """Admin Products Management Page - শুধু active products দেখাবে"""
    # 🔥 শুধু is_active=True প্রোডাক্ট দেখান (Admin যারা manually যোগ করেছে)
    products = Product.objects.filter(is_active=True).order_by('-created_at')
    
    context = {
        'products': products,
        'products_json': json.dumps([{
            'id': p.id,
            'name': p.name,
            'description': p.description,
            'price': float(p.price),
            'discount_price': float(p.discount_price) if p.discount_price else None,
            'discount_percentage': p.discount_percentage if p.discount_percentage else None,
            'stock_quantity': p.stock_quantity if hasattr(p, 'stock_quantity') else 0,
            'is_active': p.is_active,
            'special_offer': p.special_offer if hasattr(p, 'special_offer') else None,
            'image_url': p.image.url if p.image else None,
            'created_at': p.created_at.isoformat() if p.created_at else None,
        } for p in products]),
        'total_products': products.count(),
        'active_products': products.filter(is_active=True).count(),
        'inactive_products': Product.objects.filter(is_active=False).count(),
        'low_stock_products': products.filter(stock_quantity__lte=10, stock_quantity__gt=0).count() if hasattr(Product, 'stock_quantity') else 0,
        'total_orders': Order.objects.count(),
        'total_customers': 0,
    }
    return render(request, 'admin_products.html', context)
# ===== Custom Admin Edit Views =====

@csrf_exempt
@staff_member_required
def update_order(request, order_id):
    """Update an existing order via API"""
    if request.method not in ['POST', 'PUT']:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        order = get_object_or_404(Order, order_id=order_id)
        
        # Get form data
        customer_name = request.POST.get('customer_name', '').strip()
        customer_email = request.POST.get('customer_email', '').strip()
        customer_phone = request.POST.get('customer_phone', '').strip()
        customer_address = request.POST.get('customer_address', '').strip()
        customer_district = request.POST.get('customer_district', '').strip()
        status = request.POST.get('status', 'pending')
        payment_status = request.POST.get('payment_status', 'pending')
        
        # Validate required fields
        if not customer_name:
            return JsonResponse({'status': 'error', 'message': 'Customer name is required'}, status=400)
        if not customer_email:
            return JsonResponse({'status': 'error', 'message': 'Customer email is required'}, status=400)
        
        # Update order
        order.customer_name = customer_name
        order.customer_email = customer_email
        order.customer_phone = customer_phone if customer_phone else None
        order.customer_address = customer_address if customer_address else None
        order.customer_district = customer_district if customer_district else None
        order.status = status
        order.payment_status = payment_status
        
        order.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Order updated successfully',
            'order_id': order.order_id
        })
        
    except Exception as e:
        print(f"❌ Error updating order: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@staff_member_required
def admin_product_edit(request, product_id):
    """Custom Product Edit Page"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        # Update product logic here
        product.name = request.POST.get('name', product.name)
        product.description = request.POST.get('description', product.description)
        product.price = request.POST.get('price', product.price)
        product.discount_price = request.POST.get('discount_price', product.discount_price)
        product.discount_percentage = request.POST.get('discount_percentage', product.discount_percentage)
        product.stock_quantity = request.POST.get('stock_quantity', product.stock_quantity)
        product.is_active = request.POST.get('is_active', 'on') == 'on'
        product.special_offer = request.POST.get('special_offer', product.special_offer)
        product.save()
        
        messages.success(request, f'Product {product.name} updated successfully!')
        return redirect('admin_products')
    
    context = {
        'product': product,
    }
    return render(request, 'admin_product_edit.html', context)


@staff_member_required
def get_products(request):
    """Get products as JSON"""
    products = Product.objects.all().order_by('-created_at')
    data = [{
        'id': p.id,
        'name': p.name,
        'description': p.description,
        'price': float(p.price),
        'discount_price': float(p.discount_price) if p.discount_price else None,
        'discount_percentage': p.discount_percentage if p.discount_percentage else None,
        'stock_quantity': p.stock_quantity if hasattr(p, 'stock_quantity') else 0,
        'is_active': p.is_active,
        'special_offer': p.special_offer if hasattr(p, 'special_offer') else None,
        'image_url': p.image.url if p.image else None,
        'created_at': p.created_at.isoformat() if p.created_at else None,
    } for p in products]
    return JsonResponse(data, safe=False)


@staff_member_required
def get_product_detail(request, product_id):
    """Get single product detail as JSON"""
    try:
        product = get_object_or_404(Product, id=product_id)
        data = {
            'id': product.id,
            'name': product.name,
            'description': product.description,
            'price': float(product.price),
            'discount_price': float(product.discount_price) if product.discount_price else None,
            'discount_percentage': product.discount_percentage if product.discount_percentage else None,
            'stock_quantity': product.stock_quantity if hasattr(product, 'stock_quantity') else 0,
            'is_active': product.is_active,
            'special_offer': product.special_offer if hasattr(product, 'special_offer') else None,
            'image_url': product.image.url if product.image else None,
            'created_at': product.created_at.isoformat() if product.created_at else None,
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# views.py - create_product ফাংশন (Inventory চেক সহ)

# views.py - create_product ফাংশন

# views.py - create_product ফাংশন (সঠিকভাবে inventory কমাবে)

@csrf_exempt
@staff_member_required
def create_product(request):
    """Create a new product - Inventory stock REDUCES, Stock Quantity INCREASES"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get form data
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        price = request.POST.get('price')
        discount_price = request.POST.get('discount_price', '').strip()
        discount_percentage = request.POST.get('discount_percentage', '').strip()
        stock_quantity = request.POST.get('stock_quantity', 0)
        is_active = request.POST.get('is_active', 'true') == 'true'
        special_offer = request.POST.get('special_offer', '').strip()
        
        # Validate
        if not name:
            return JsonResponse({'status': 'error', 'message': 'Product name is required'}, status=400)
        if not price:
            return JsonResponse({'status': 'error', 'message': 'Price is required'}, status=400)
        
        try:
            price = float(price)
            if price < 0:
                return JsonResponse({'status': 'error', 'message': 'Price cannot be negative'}, status=400)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid price format'}, status=400)
        
        try:
            stock_quantity = int(stock_quantity)
            if stock_quantity < 0:
                return JsonResponse({'status': 'error', 'message': 'Stock cannot be negative'}, status=400)
        except ValueError:
            stock_quantity = 0
        
        # Check discount
        is_on_sale = False
        discount_price_value = None
        discount_percentage_value = None
        
        if discount_price and discount_price != '':
            try:
                discount_price_value = float(discount_price)
                if discount_price_value > 0 and discount_price_value < price:
                    is_on_sale = True
            except:
                pass
        
        if not is_on_sale and discount_percentage and discount_percentage != '':
            try:
                discount_percentage_value = int(discount_percentage)
                if discount_percentage_value > 0 and discount_percentage_value <= 100:
                    is_on_sale = True
            except:
                pass
        
        # ============================================================
        # 🔥 STEP 1: INVENTORY CHECK
        # ============================================================
        if stock_quantity > 0:
            from water_app.models import Inventory
            from django.db import models
            
            # Get total inventory available
            inventory_total = Inventory.objects.aggregate(
                total=models.Sum('quantity')
            )['total'] or 0
            
            print(f"📊 Total inventory available: {inventory_total}")
            print(f"📊 Requested stock: {stock_quantity}")
            
            # Check if sufficient inventory
            if inventory_total < stock_quantity:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Insufficient inventory. Available: {inventory_total}, Required: {stock_quantity}'
                }, status=400)
            
            print(f"✅ Sufficient inventory available")
        
        # ============================================================
        # 🔥 STEP 2: CREATE PRODUCT
        # ============================================================
        product = Product(
            name=name,
            description=description,
            price=price,
            discount_price=discount_price_value,
            discount_percentage=discount_percentage_value,
            stock_quantity=stock_quantity,  # ✅ Stock Quantity বাড়বে
            is_active=is_active,
            special_offer=special_offer if special_offer else None,
            is_on_sale=is_on_sale,
        )
        
        # Handle image
        if 'image' in request.FILES:
            image_file = request.FILES['image']
            if image_file.content_type not in ['image/jpeg', 'image/png', 'image/gif', 'image/webp']:
                return JsonResponse({'status': 'error', 'message': 'Invalid image format'}, status=400)
            if image_file.size > 2 * 1024 * 1024:
                return JsonResponse({'status': 'error', 'message': 'Image size must be less than 2MB'}, status=400)
            product.image = image_file
        
        product.save()
        
        # ============================================================
        # 🔥 STEP 3: REDUCE INVENTORY (Inventory Stock কমবে)
        # ============================================================
        if stock_quantity > 0:
            from water_app.models import Inventory
            from django.db import models
            
            # Get current inventory total
            inventory_total = Inventory.objects.aggregate(
                total=models.Sum('quantity')
            )['total'] or 0
            
            # 🔥 IMPORTANT: Inventory কমাবে
            inventory = Inventory.objects.create(
                product=product,
                quantity=-stock_quantity,  # 🔥 Negative: Inventory কমবে
                previous_stock=inventory_total,
                new_stock=inventory_total - stock_quantity,
                movement_type='adjustment',
                reference=f'Product Created: {product.id}',
                notes=f'Stock TAKEN from inventory for product: {product.name}. Quantity: {stock_quantity}',
                user=request.user if request.user.is_authenticated else None
            )
            
            print(f"✅ Inventory REDUCED by {stock_quantity}")
            print(f"   New inventory total: {inventory_total - stock_quantity}")
            print(f"   Product stock_quantity: {product.stock_quantity}")
        
        return JsonResponse({
            'status': 'success',
            'message': f'Product "{name}" created. Stock: +{stock_quantity}, Inventory: -{stock_quantity}',
            'product_id': product.id,
            'stock_quantity': product.stock_quantity,
            'inventory_reduced': stock_quantity
        })
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)    

@csrf_exempt
@staff_member_required
def update_product(request, product_id):
    """Update an existing product - Inventory adjusts accordingly"""
    if request.method not in ['PUT', 'POST']:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        product = get_object_or_404(Product, id=product_id)
        
        # Get form data
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        price = request.POST.get('price')
        discount_price = request.POST.get('discount_price', '').strip()
        discount_percentage = request.POST.get('discount_percentage', '').strip()
        stock_quantity = request.POST.get('stock_quantity', 0)
        is_active = request.POST.get('is_active', 'true') == 'true'
        special_offer = request.POST.get('special_offer', '').strip()
        
        # Validate required fields
        if not name:
            return JsonResponse({'status': 'error', 'message': 'Product name is required'}, status=400)
        if not price:
            return JsonResponse({'status': 'error', 'message': 'Price is required'}, status=400)
        
        try:
            price = float(price)
            if price < 0:
                return JsonResponse({'status': 'error', 'message': 'Price cannot be negative'}, status=400)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid price format'}, status=400)
        
        try:
            stock_quantity = int(stock_quantity)
            if stock_quantity < 0:
                return JsonResponse({'status': 'error', 'message': 'Stock cannot be negative'}, status=400)
        except ValueError:
            stock_quantity = 0
        
        # Store old stock
        old_stock = product.stock_quantity
        
        # Check if product has discount
        is_on_sale = False
        discount_price_value = None
        discount_percentage_value = None
        
        if discount_price and discount_price != '':
            try:
                discount_price_value = float(discount_price)
                if discount_price_value > 0 and discount_price_value < price:
                    is_on_sale = True
            except:
                pass
        
        if not is_on_sale and discount_percentage and discount_percentage != '':
            try:
                discount_percentage_value = int(discount_percentage)
                if discount_percentage_value > 0 and discount_percentage_value <= 100:
                    is_on_sale = True
            except:
                pass
        
        # Determine quantity difference before saving
        quantity_diff = stock_quantity - old_stock

        # If product stock would increase, ensure inventory has enough BEFORE saving
        if quantity_diff > 0:
            try:
                from water_app.models import Inventory
                inventory_total_check = Inventory.objects.filter(product=product).aggregate(
                    total=models.Sum('quantity')
                )['total'] or 0
                if inventory_total_check < quantity_diff:
                    return JsonResponse({'status': 'error', 'message': f'Insufficient inventory. Available inventory: {inventory_total_check}'}, status=400)
            except Exception as e:
                print(f"⚠️ Inventory check skipped: {str(e)}")

        # Update product
        product.name = name
        product.description = description
        product.price = price
        product.discount_price = discount_price_value
        product.discount_percentage = discount_percentage_value
        product.stock_quantity = stock_quantity
        product.is_active = is_active
        product.special_offer = special_offer if special_offer else None
        product.is_on_sale = is_on_sale
        
        # Handle image upload
        if 'image' in request.FILES:
            if product.image:
                try:
                    default_storage.delete(product.image.path)
                except:
                    pass
            product.image = request.FILES['image']
        elif 'remove_image' in request.POST and request.POST.get('remove_image') == 'true':
            if product.image:
                try:
                    default_storage.delete(product.image.path)
                except:
                    pass
                product.image = None
        
        product.save()
        
        # 👇 Inventory adjusts when product stock changes
        if stock_quantity != old_stock:
            try:
                from water_app.models import Inventory
                quantity_diff = stock_quantity - old_stock

                # Get current inventory total
                inventory_total = Inventory.objects.filter(product=product).aggregate(
                    total=models.Sum('quantity')
                )['total'] or 0

                # If product stock increases (quantity_diff > 0), ensure inventory has enough
                if quantity_diff > 0 and inventory_total < quantity_diff:
                    return JsonResponse({'status': 'error', 'message': f'Insufficient inventory. Available inventory: {inventory_total}'}, status=400)

                # Inventory changes opposite to product stock
                # If product stock increases, inventory decreases (stock taken)
                # If product stock decreases, inventory increases (stock returned)
                inventory_change = -quantity_diff

                Inventory.objects.create(
                    product=product,
                    quantity=inventory_change,
                    previous_stock=inventory_total,
                    new_stock=inventory_total + inventory_change,
                    movement_type='adjustment',
                    reference=f'Product Updated: {product.id}',
                    notes=f'Stock changed from {old_stock} to {stock_quantity}. Inventory adjusted by {inventory_change}',
                    user=request.user if request.user.is_authenticated else None
                )

                print(f"✅ Inventory adjusted: {product.name} - Change: {inventory_change}")

            except Exception as e:
                print(f"⚠️ Inventory update skipped: {str(e)}")
        
        return JsonResponse({
            'status': 'success',
            'message': f'Product "{name}" updated successfully.',
            'product_id': product.id
        })
        
    except Exception as e:
        print(f"❌ Error updating product: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
@csrf_exempt
@staff_member_required
def delete_product(request, product_id):
    """Delete a product"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        product = get_object_or_404(Product, id=product_id)
        
        # Delete image if exists
        if product.image:
            try:
                default_storage.delete(product.image.path)
            except:
                pass
        
        product.delete()
        return JsonResponse({'status': 'success', 'message': 'Product deleted successfully'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@staff_member_required
def toggle_product_status(request, product_id):
    """Toggle product active status"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        product = get_object_or_404(Product, id=product_id)
        data = json.loads(request.body)
        is_active = data.get('is_active', not product.is_active)
        
        product.is_active = is_active
        product.save()
        
        return JsonResponse({
            'status': 'success',
            'message': f'Product {"activated" if is_active else "deactivated"} successfully',
            'is_active': is_active
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@staff_member_required
def export_products_excel(request):
    """Export products as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse("⚠️ openpyxl not installed. Please install: pip install openpyxl", status=500)
    
    products = Product.objects.all().order_by('-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Report"
    
    # Styles
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = "📦 Aquanimity SuperWater - Products Report"
    title_cell.font = Font(name='Arial', size=18, bold=True, color='1a5276')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:J2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total Products: {products.count()}"
    subtitle_cell.font = Font(name='Arial', size=10, color='666666')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Sl No', 'Product Name', 'Description', 'Price', 'Discount Price', 'Discount %', 'Stock', 'Status', 'Special Offer', 'Created At']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    row_num = 5
    for idx, product in enumerate(products, 1):
        row_data = [
            idx,
            product.name,
            product.description[:50] + ('...' if product.description and len(product.description) > 50 else '') or '-',
            f"৳{float(product.price):.2f}",
            f"৳{float(product.discount_price):.2f}" if product.discount_price else '-',
            f"{product.discount_percentage}%" if product.discount_percentage else '-',
            product.stock_quantity if hasattr(product, 'stock_quantity') else 0,
            'Active' if product.is_active else 'Inactive',
            product.special_offer if hasattr(product, 'special_offer') and product.special_offer else '-',
            product.created_at.strftime('%Y-%m-%d %H:%M') if product.created_at else '-'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
        
        row_num += 1
    
    # Column widths
    widths = {'A': 6, 'B': 30, 'C': 35, 'D': 12, 'E': 14, 'F': 10, 'G': 10, 'H': 10, 'I': 25, 'J': 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A5'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'products_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def export_products_pdf(request):
    """Export products as PDF file"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return HttpResponse("⚠️ ReportLab not installed. Please install: pip install reportlab", status=500)
    
    products = Product.objects.all().order_by('-created_at')
    
    response = HttpResponse(content_type='application/pdf')
    filename = f'products_report_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a5276'), alignment=TA_CENTER, spaceAfter=12, fontName='Helvetica-Bold')
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'), alignment=TA_CENTER, spaceAfter=16)
    
    story = []
    story.append(Paragraph("📦 Aquanimity SuperWater - Products Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    story.append(Paragraph(f"Total Products: {products.count()}", subtitle_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table
    table_data = [['Sl', 'Product Name', 'Price', 'Discount Price', 'Discount %', 'Stock', 'Status']]
    for idx, product in enumerate(products[:100], 1):
        table_data.append([
            str(idx),
            product.name[:25] + ('...' if len(product.name) > 25 else ''),
            f"৳{float(product.price):.2f}",
            f"৳{float(product.discount_price):.2f}" if product.discount_price else '-',
            f"{product.discount_percentage}%" if product.discount_percentage else '-',
            str(product.stock_quantity if hasattr(product, 'stock_quantity') else 0),
            'Active' if product.is_active else 'Inactive'
        ])
    
    if products.count() > 100:
        table_data.append(['...', f'... and {products.count() - 100} more products', '', '', '', '', ''])
    
    table = Table(table_data, colWidths=[0.4*inch, 2.0*inch, 0.8*inch, 0.9*inch, 0.7*inch, 0.6*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ddd')),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
        ('PADDING', (0, 0), (-1, -1), 3),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle('FooterStyle', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#999999'), alignment=TA_CENTER)
    story.append(Paragraph("© 2026 Aquanimity Super Water. All rights reserved.", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response


# ============================================
# DELIVERY MANAGEMENT VIEWS
# ============================================


@staff_member_required
def admin_delivery(request):
    """Admin Delivery Charges Management Page"""
    districts = DistrictDeliveryCharge.objects.all().order_by('district_name')
    
    context = {
        'districts': districts,
        'districts_json': json.dumps([{
            'id': d.id,
            'district': d.district,
            'district_name': d.district_name,
            'charge': float(d.charge),
            'delivery_time': d.delivery_time,
            'is_active': d.is_active,
            'created_at': d.created_at.isoformat() if d.created_at else None,
        } for d in districts]),
        'total_districts': districts.count(),
        'active_districts': districts.filter(is_active=True).count(),
        'inactive_districts': districts.filter(is_active=False).count(),
        'avg_charge': districts.aggregate(avg_charge=models.Avg('charge'))['avg_charge'] or 0,
        'total_orders': Order.objects.count(),
        'total_products': Product.objects.count(),
        'total_customers': 0,
    }
    return render(request, 'admin_delivery.html', context)


@csrf_exempt
@staff_member_required
def create_district(request):
    """Create a new district delivery charge"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        district_name = request.POST.get('district_name', '').strip()
        district = request.POST.get('district', '').strip().lower()
        charge = request.POST.get('charge')
        delivery_time = request.POST.get('delivery_time', '2-3 days').strip()
        is_active = request.POST.get('is_active', 'true') == 'true'
        
        # Validate required fields
        if not district_name:
            return JsonResponse({'status': 'error', 'message': 'District name is required'}, status=400)
        if not charge:
            return JsonResponse({'status': 'error', 'message': 'Delivery charge is required'}, status=400)
        
        try:
            charge = float(charge)
            if charge < 0:
                return JsonResponse({'status': 'error', 'message': 'Charge cannot be negative'}, status=400)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid charge format'}, status=400)
        
        # Auto-generate district code if not provided
        if not district:
            district = district_name.lower().replace(' ', '_')
        
        # Check if district already exists
        if DistrictDeliveryCharge.objects.filter(district=district).exists():
            return JsonResponse({'status': 'error', 'message': f'District "{district}" already exists'}, status=400)
        
        # Create district
        district_obj = DistrictDeliveryCharge(
            district_name=district_name,
            district=district,
            charge=charge,
            delivery_time=delivery_time,
            is_active=is_active,
        )
        district_obj.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'District added successfully',
            'district_id': district_obj.id
        })
        
    except Exception as e:
        print(f"Error creating district: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@staff_member_required
def update_district(request, district_id):
    """Update an existing district delivery charge"""
    if request.method not in ['POST', 'PUT']:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        district_obj = get_object_or_404(DistrictDeliveryCharge, id=district_id)
        
        district_name = request.POST.get('district_name', '').strip()
        district = request.POST.get('district', '').strip().lower()
        charge = request.POST.get('charge')
        delivery_time = request.POST.get('delivery_time', '2-3 days').strip()
        is_active = request.POST.get('is_active', 'true') == 'true'
        
        # Validate required fields
        if not district_name:
            return JsonResponse({'status': 'error', 'message': 'District name is required'}, status=400)
        if not charge:
            return JsonResponse({'status': 'error', 'message': 'Delivery charge is required'}, status=400)
        
        try:
            charge = float(charge)
            if charge < 0:
                return JsonResponse({'status': 'error', 'message': 'Charge cannot be negative'}, status=400)
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid charge format'}, status=400)
        
        # Check if district code already exists (excluding current)
        if district and DistrictDeliveryCharge.objects.filter(district=district).exclude(id=district_id).exists():
            return JsonResponse({'status': 'error', 'message': f'District "{district}" already exists'}, status=400)
        
        # Update district
        district_obj.district_name = district_name
        district_obj.district = district if district else district_name.lower().replace(' ', '_')
        district_obj.charge = charge
        district_obj.delivery_time = delivery_time
        district_obj.is_active = is_active
        district_obj.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'District updated successfully',
            'district_id': district_obj.id
        })
        
    except Exception as e:
        print(f"Error updating district: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@staff_member_required
def delete_district(request, district_id):
    """Delete a district delivery charge"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        district_obj = get_object_or_404(DistrictDeliveryCharge, id=district_id)
        district_obj.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': 'District deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@staff_member_required
def toggle_district_status(request, district_id):
    """Toggle district active status"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        district_obj = get_object_or_404(DistrictDeliveryCharge, id=district_id)
        data = json.loads(request.body)
        is_active = data.get('is_active', not district_obj.is_active)
        
        district_obj.is_active = is_active
        district_obj.save()
        
        return JsonResponse({
            'status': 'success',
            'message': f'District {"activated" if is_active else "deactivated"} successfully',
            'is_active': is_active
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@staff_member_required
def export_districts_excel(request):
    """Export districts as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse("⚠️ openpyxl not installed. Please install: pip install openpyxl", status=500)
    
    districts = DistrictDeliveryCharge.objects.all().order_by('district_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Districts Report"
    
    # Styles
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a5276', end_color='1a5276', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "📦 Aquanimity SuperWater - Delivery Charges Report"
    title_cell.font = Font(name='Arial', size=18, bold=True, color='1a5276')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')} | Total Districts: {districts.count()}"
    subtitle_cell.font = Font(name='Arial', size=10, color='666666')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Sl No', 'District Name', 'District Code', 'Charge (৳)', 'Delivery Time', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    row_num = 5
    for idx, district in enumerate(districts, 1):
        row_data = [
            idx,
            district.district_name,
            district.district,
            float(district.charge),
            district.delivery_time or '2-3 days',
            'Active' if district.is_active else 'Inactive'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
        
        row_num += 1
    
    # Column widths
    widths = {'A': 6, 'B': 30, 'C': 18, 'D': 14, 'E': 16, 'F': 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A5'
    
    # Summary Statistics
    summary_row = row_num + 2
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    summary_title = ws.cell(row=summary_row, column=1)
    summary_title.value = "📊 Summary Statistics"
    summary_title.font = Font(name='Arial', size=12, bold=True)
    
    summary_data = [
        ['Total Districts', str(districts.count())],
        ['Active Districts', str(districts.filter(is_active=True).count())],
        ['Inactive Districts', str(districts.filter(is_active=False).count())],
        ['Average Charge', f"৳ {districts.aggregate(avg_charge=models.Avg('charge'))['avg_charge'] or 0:.2f}"],
    ]
    
    for idx, (label, value) in enumerate(summary_data):
        row = summary_row + idx + 1
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
    
    # Footer
    footer_row = summary_row + len(summary_data) + 2
    ws.merge_cells(f'A{footer_row}:F{footer_row}')
    footer_cell = ws.cell(row=footer_row, column=1)
    footer_cell.value = "© 2026 Aquanimity Super Water. All rights reserved."
    footer_cell.font = Font(name='Arial', size=9, color='999999')
    footer_cell.alignment = Alignment(horizontal='center')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'districts_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@staff_member_required
def export_districts_pdf(request):
    """Export districts as PDF file"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return HttpResponse("⚠️ ReportLab not installed. Please install: pip install reportlab", status=500)
    
    districts = DistrictDeliveryCharge.objects.all().order_by('district_name')
    
    response = HttpResponse(content_type='application/pdf')
    filename = f'districts_report_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a5276'), alignment=TA_CENTER, spaceAfter=12, fontName='Helvetica-Bold')
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'), alignment=TA_CENTER, spaceAfter=16)
    
    story = []
    story.append(Paragraph("📦 Aquanimity SuperWater - Delivery Charges Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    story.append(Paragraph(f"Total Districts: {districts.count()}", subtitle_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table
    table_data = [['Sl', 'District Name', 'District Code', 'Charge (৳)', 'Delivery Time', 'Status']]
    for idx, district in enumerate(districts[:100], 1):
        table_data.append([
            str(idx),
            district.district_name[:25] + ('...' if len(district.district_name) > 25 else ''),
            district.district,
            f"৳{float(district.charge):.2f}",
            district.delivery_time or '2-3 days',
            'Active' if district.is_active else 'Inactive'
        ])
    
    if districts.count() > 100:
        table_data.append(['...', f'... and {districts.count() - 100} more districts', '', '', '', ''])
    
    table = Table(table_data, colWidths=[0.4*inch, 2.2*inch, 1.0*inch, 0.8*inch, 0.8*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ddd')),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ('PADDING', (0, 0), (-1, -1), 3),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle('FooterStyle', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#999999'), alignment=TA_CENTER)
    story.append(Paragraph("© 2026 Aquanimity Super Water. All rights reserved.", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response

# views.py - Inventory Management Section (Cleaned & Updated)

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import Product, Inventory
import json
import csv
from datetime import datetime
from decimal import Decimal


# ============================================================
# 🔥 ADMIN INVENTORY VIEW
# ============================================================
# views.py - admin_inventory view আপডেট করুন
# views.py - admin_inventory আপডেট করুন

@login_required
def admin_inventory(request):
    """Admin inventory management view"""
    if not request.user.is_staff and not request.user.is_superuser:
        return render(request, '403.html', {'message': 'Admin access required'}, status=403)
    
    # 🔥 সব প্রোডাক্ট দেখান (active + inventory থাকা)
    # অথবা শুধু যাদের inventory আছে তাদের দেখান
    from django.db import models
    from .models import Inventory
    
    # 🔥 শুধু যাদের inventory আছে তাদের দেখান
    products_with_inventory = Product.objects.filter(
        models.Exists(Inventory.objects.filter(product=models.OuterRef('pk')))
    ).order_by('-created_at')
    
    # অথবা সব প্রোডাক্ট দেখাতে চাইলে:
    # products = Product.objects.all().order_by('-created_at')
    
    products_data = []
    for product in products_with_inventory:
        try:
            inventory_total = product.get_inventory_total()
        except:
            inventory_total = 0
        
        stock_quantity = product.stock_quantity or 0
        total_available = inventory_total + stock_quantity
        
        image_url = None
        if product.image and hasattr(product.image, 'url'):
            image_url = product.image.url
        
        product_data = {
            'id': product.id,
            'name': product.name,
            'price': product.price,
            'category': product.category,
            'description': product.description,
            'image_url': image_url,
            'is_on_sale': product.is_on_sale,
            'discount_price': product.discount_price,
            'discount_percentage': product.discount_percentage,
            'special_offer': product.special_offer,
            'created_at': product.created_at,
            'stock_quantity': stock_quantity,
            'inventory_total': inventory_total,
            'total_available': total_available,
            'is_active': product.is_active,  # 🔥 active status দেখান
        }
        products_data.append(product_data)
    
    # Calculate stats
    total_products = len(products_data)
    in_stock = sum(1 for p in products_data if p['inventory_total'] > 20)
    low_stock = sum(1 for p in products_data if 0 < p['inventory_total'] <= 20)
    out_of_stock = sum(1 for p in products_data if p['inventory_total'] == 0)
    
    # JSON conversion
    products_json = []
    for product in products_data:
        products_json.append({
            'id': product['id'],
            'name': product['name'],
            'price': float(product['price']),
            'stock_quantity': product['stock_quantity'],
            'inventory_total': product['inventory_total'],
            'total_available': product['total_available'],
            'category': product['category'],
            'description': product['description'] or '',
            'image_url': product['image_url'],
            'is_on_sale': product['is_on_sale'],
            'discount_price': float(product['discount_price']) if product['discount_price'] else None,
            'discount_percentage': product['discount_percentage'],
            'special_offer': product['special_offer'] or '',
            'created_at': product['created_at'].strftime('%Y-%m-%d %H:%M:%S') if product['created_at'] else None,
            'is_active': product['is_active'],
        })
    
    context = {
        'products': products_data,
        'products_json': json.dumps(products_json),
        'total_products': total_products,
        'in_stock_products': in_stock,
        'low_stock_products': low_stock,
        'out_of_stock_products': out_of_stock,
    }
    
    return render(request, 'admin_inventory.html', context)
# ============================================================
# 🔥 DELETE PRODUCT FROM INVENTORY (শুধু ইনভেন্টরি থেকে ডিলিট)
# ============================================================

@csrf_exempt
@require_http_methods(["DELETE"])
@login_required
def delete_inventory_product(request, product_id):
    """
    🔥 admin_inventory থেকে প্রোডাক্ট ডিলিট করলে:
    - শুধু ইনভেন্টরি রেকর্ড ডিলিট হবে
    - প্রোডাক্ট ডিলিট হবে না (admin_products এ থাকবে)
    """
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        product = get_object_or_404(Product, id=product_id)
        
        # 🔥 শুধু ইনভেন্টরি রেকর্ড ডিলিট করুন
        Inventory.objects.filter(product=product).delete()
        
        # অথবা আপনি চাইলে stock_quantity রিসেট করতে পারেন
        product.stock_quantity = 0
        product.save()
        
        return JsonResponse({
            'status': 'success',
            'message': f'Inventory records deleted for {product.name}'
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


# ============================================================
# 🔥 EDIT INVENTORY PRODUCT (শুধু ইনভেন্টরি এডিট)
# ============================================================

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def edit_inventory_product(request, product_id):
    """
    🔥 admin_inventory থেকে প্রোডাক্ট এডিট করলে:
    - শুধু ইনভেন্টরি আপডেট হবে
    - প্রোডাক্টের price ইত্যাদি admin_products এ আপডেট করতে হবে
    """
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        product = get_object_or_404(Product, id=product_id)
        data = json.loads(request.body)
        
        # 🔥 শুধু ইনভেন্টরি সম্পর্কিত ফিল্ড আপডেট করুন
        # (price, name ইত্যাদি admin_products এ আপডেট হবে)
        
        # আপনি চাইলে ইনভেন্টরি এডিট করতে পারেন
        inventory_quantity = data.get('inventory_quantity')
        if inventory_quantity is not None:
            current_inventory = product.get_inventory_total()
            diff = inventory_quantity - current_inventory
            
            if diff != 0:
                Inventory.objects.create(
                    product=product,
                    quantity=diff,
                    previous_stock=current_inventory,
                    new_stock=inventory_quantity,
                    movement_type='adjustment',
                    notes=f'Inventory adjusted from {current_inventory} to {inventory_quantity}',
                    user=request.user
                )
        
        return JsonResponse({
            'status': 'success',
            'message': f'Inventory updated for {product.name}'
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
# ============================================================
# 🔥 1. ADMIN_INVENTORY থেকে যোগ করলে - ইনভেন্টরি বাড়বে
# ============================================================

# views.py - add_product_by_name_api

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def add_product_by_name_api(request):
    """
    🔥 admin_inventory থেকে প্রোডাক্ট যোগ করলে:
    - is_active = False হবে
    - admin_products এ দেখাবে না
    """
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        data = json.loads(request.body)
        product_name = data.get('name', '').strip()
        quantity = data.get('quantity', 0)
        price = data.get('price', 0)
        description = data.get('description', '')
        
        if not product_name:
            return JsonResponse({'status': 'error', 'message': 'Product name is required'}, status=400)
        
        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': 'Quantity must be greater than 0'}, status=400)
        
        # চেক করুন প্রোডাক্ট ইতিমধ্যে আছে কিনা
        existing_product = Product.objects.filter(name__iexact=product_name).first()
        
        if existing_product:
            # প্রোডাক্ট থাকলে শুধু ইনভেন্টরি বাড়বে
            current_inventory = existing_product.get_inventory_total()
            new_inventory = current_inventory + quantity
            
            Inventory.objects.create(
                product=existing_product,
                quantity=quantity,
                previous_stock=current_inventory,
                new_stock=new_inventory,
                movement_type='add',
                notes=f'Added {quantity} units from admin_inventory - Added by {request.user.email}',
                user=request.user
            )
            
            return JsonResponse({
                'status': 'success',
                'message': f'Added {quantity} units to inventory: {product_name}',
                'product_id': existing_product.id,
                'product_name': existing_product.name,
                'inventory_total': new_inventory,
                'stock_quantity': existing_product.stock_quantity,
                'total_available': new_inventory + existing_product.stock_quantity,
                'action': 'inventory_added'
            })
        else:
            # 🔥 নতুন প্রোডাক্ট তৈরি করুন - is_active=False সহ
            product = Product(
                name=product_name,
                price=Decimal(str(price)),
                stock_quantity=0,
                description=description,
                category='Filters',
                # 🔥 is_active False সেট করুন
                is_active=False
            )
            # জোর করে False সেট করুন
            product.is_active = False
            product.save()
            
            print(f"✅ New product created: {product_name} (is_active={product.is_active})")
            
            # Inventory তৈরি করুন
            Inventory.objects.create(
                product=product,
                quantity=quantity,
                previous_stock=0,
                new_stock=quantity,
                movement_type='product_created',
                notes=f'Product created with {quantity} units from admin_inventory (inactive in admin_products)',
                user=request.user
            )
            
            return JsonResponse({
                'status': 'success',
                'message': f'Product "{product_name}" added to inventory with {quantity} units! (Hidden from admin_products)',
                'product_id': product.id,
                'product_name': product.name,
                'inventory_total': quantity,
                'stock_quantity': 0,
                'total_available': quantity,
                'action': 'product_created_inventory_only',
                'warning': 'This product is hidden from admin_products (is_active=False)'
            })
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# ============================================================
# 🔥 2. ADMIN_PRODUCTS থেকে যোগ করলে - স্টক বাড়বে, ইনভেন্টরি কমবে
# ============================================================



# ============================================================
# 🔥 3. ADD STOCK API (admin_inventory থেকে স্টক যোগ)
# ============================================================
# views.py
# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import Product, Inventory
import json
# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import Product, Inventory
import json

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def add_stock_api(request, product_id=None):
    """
    API: Add stock to existing product - Admin only
    Supports both:
    - /api/inventory/add-stock/ (product_id in body)
    - /api/inventory/<int:product_id>/add-stock/ (product_id in URL)
    """
    try:
        print("=" * 50)
        print("🔵 ADD STOCK API CALLED")
        print(f"Product ID from URL: {product_id}")
        
        # Check admin access
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        # Get data
        data = json.loads(request.body)
        print(f"📥 Request data: {data}")
        
        # Get product_id from URL or body
        if product_id:
            pid = product_id
            quantity = data.get('quantity', 0)
        else:
            pid = data.get('product_id')
            quantity = data.get('quantity', 0)
        
        print(f"📦 Product ID: {pid}")
        print(f"📦 Quantity: {quantity}")
        
        # Validation
        if not pid:
            return JsonResponse({
                'status': 'error',
                'message': 'Product ID required'
            }, status=400)
        
        if quantity <= 0:
            return JsonResponse({
                'status': 'error',
                'message': 'Quantity must be greater than 0'
            }, status=400)
        
        # Get product
        try:
            product = Product.objects.get(id=pid)
            print(f"✅ Product found: {product.name}")
        except Product.DoesNotExist:
            print(f"❌ Product not found: {pid}")
            return JsonResponse({
                'status': 'error',
                'message': 'Product not found'
            }, status=404)
        
        # Get current inventory
        try:
            current_inventory = product.get_inventory_total()
        except:
            current_inventory = product.stock_quantity or 0
        
        new_inventory = current_inventory + quantity
        print(f"📊 Current: {current_inventory}, New: {new_inventory}")
        
        # Create inventory entry
        try:
            inventory = Inventory.objects.create(
                product=product,
                quantity=quantity,
                previous_stock=current_inventory,
                new_stock=new_inventory,
                movement_type='restock',
                notes=f'Added {quantity} units - Added by {request.user.email}',
                user=request.user
            )
            print(f"✅ Inventory entry created: {inventory.id}")
        except Exception as e:
            print(f"⚠️ Inventory entry failed: {e}")
            # Fallback: update stock_quantity directly
            product.stock_quantity = new_inventory
            product.save()
            print(f"✅ Product stock_quantity updated directly")
        
        return JsonResponse({
            'status': 'success',
            'message': f'Added {quantity} units to {product.name}',
            'product_id': product.id,
            'product_name': product.name,
            'inventory_total': new_inventory,
            'stock_quantity': product.stock_quantity,
            'total_available': new_inventory + (product.stock_quantity or 0),
            'previous_stock': current_inventory,
            'new_stock': new_inventory
        })
        
    except json.JSONDecodeError as e:
        print(f"❌ JSON Decode Error: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': f'Invalid JSON data: {str(e)}'
        }, status=400)
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
    

# ============================================================
# 🔥 4. EXPORT INVENTORY EXCEL
# ============================================================

@login_required
def export_inventory_excel(request):
    """Export inventory data as CSV/Excel"""
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Admin access required'}, status=403)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Product ID', 'Product Name', 'Category', 'Price', 
            'Stock Quantity', 'Inventory Total', 'Total Available',
            'Inventory Value', 'Total Stock Value', 'Status', 'Created At'
        ])
        
        products = Product.objects.filter(is_active=True).order_by('-created_at')
        
        for product in products:
            try:
                inventory_total = product.get_inventory_total()
            except:
                inventory_total = 0
            
            stock_quantity = product.stock_quantity or 0
            total_available = inventory_total + stock_quantity
            inventory_value = float(product.price) * inventory_total
            total_stock_value = float(product.price) * total_available
            
            status = 'In Stock' if total_available > 20 else 'Low Stock' if total_available > 0 else 'Out of Stock'
            
            writer.writerow([
                product.id,
                product.name,
                product.category,
                float(product.price),
                stock_quantity,
                inventory_total,
                total_available,
                inventory_value,
                total_stock_value,
                status,
                product.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ============================================================
# 🔥 5. EXPORT INVENTORY PDF
# ============================================================

@login_required
def export_inventory_pdf(request):
    """Export inventory data as PDF"""
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Admin access required'}, status=403)
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            return JsonResponse({
                'status': 'error', 
                'message': 'ReportLab not installed. Please install: pip install reportlab'
            }, status=400)
        
        import io
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#007bff'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        story = []
        story.append(Paragraph("📦 Inventory Report - Superwater", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", 
                              ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#666'), alignment=TA_CENTER)))
        story.append(Spacer(1, 0.2*inch))
        
        table_data = [['ID', 'Product', 'Category', 'Price', 'Stock', 'Inventory', 'Total', 'Value', 'Status']]
        
        products = Product.objects.filter(is_active=True).order_by('-created_at')
        
        for product in products[:100]:
            try:
                inventory_total = product.get_inventory_total()
            except:
                inventory_total = 0
            
            stock_quantity = product.stock_quantity or 0
            total_available = inventory_total + stock_quantity
            inventory_value = float(product.price) * inventory_total
            
            if total_available > 20:
                status = 'In Stock'
            elif total_available > 0:
                status = 'Low Stock'
            else:
                status = 'Out of Stock'
            
            table_data.append([
                str(product.id),
                product.name[:25] + ('...' if len(product.name) > 25 else ''),
                product.category,
                f"৳{float(product.price):.2f}",
                str(stock_quantity),
                str(inventory_total),
                str(total_available),
                f"৳{inventory_value:.2f}",
                status
            ])
        
        table = Table(table_data, colWidths=[0.4*inch, 1.5*inch, 0.8*inch, 0.7*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.8*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ddd')),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 3),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("© 2026 Superwater - All rights reserved", 
                              ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#999'), alignment=TA_CENTER)))
        
        doc.build(story)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
        
    except Exception as e:
        print(f"❌ PDF Export error: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
# ============================================
# CUSTOMER MANAGEMENT VIEWS
# ============================================

from water_app.models import Customer, Order

@staff_member_required
def admin_customers(request):
    """Admin Customers Management Page"""
    customers = Customer.objects.all().order_by('-created_at')
    
    customers_json = []
    total_revenue = 0
    total_orders = 0
    
    for customer in customers:
        # Get customer orders
        orders = Order.objects.filter(customer_email=customer.email)
        order_count = orders.count()
        total_spent = sum(float(o.total_amount) for o in orders)
        
        total_revenue += total_spent
        total_orders += order_count
        
        customers_json.append({
            'id': customer.id,
            'name': customer.name,
            'email': customer.email,
            'phone': customer.phone,
            'address': customer.address,
            'district': customer.district,
            'is_active': customer.is_active,
            'is_diabetic': customer.is_diabetic,
            'diabetes_type': customer.diabetes_type,
            'has_high_blood_pressure': customer.has_high_blood_pressure,
            'blood_pressure_status': customer.blood_pressure_status,
            'order_count': order_count,
            'total_spent': total_spent,
            'created_at': customer.created_at.isoformat() if customer.created_at else None,
        })
    
    context = {
        'active_page': 'admin_customers',
        'customers': customers,
        'customers_json': json.dumps(customers_json),
        'total_customers': customers.count(),
        'active_customers': customers.filter(is_active=True).count(),
        'inactive_customers': customers.filter(is_active=False).count(),
        'total_revenue': total_revenue,
        'total_orders': total_orders,
        'total_products': Product.objects.count(),
        'total_districts': DistrictDeliveryCharge.objects.count(),
    }
    return render(request, 'admin_customer.html', context)


@staff_member_required
def get_customers(request):
    """Get customers as JSON"""
    customers = Customer.objects.all().order_by('-created_at')
    data = []
    for customer in customers:
        orders = Order.objects.filter(customer_email=customer.email)
        data.append({
            'id': customer.id,
            'name': customer.name,
            'email': customer.email,
            'phone': customer.phone,
            'address': customer.address,
            'district': customer.district,
            'is_active': customer.is_active,
            'is_diabetic': customer.is_diabetic,
            'diabetes_type': customer.diabetes_type,
            'has_high_blood_pressure': customer.has_high_blood_pressure,
            'blood_pressure_status': customer.blood_pressure_status,
            'order_count': orders.count(),
            'total_spent': sum(float(o.total_amount) for o in orders),
            'created_at': customer.created_at.isoformat() if customer.created_at else None,
        })
    return JsonResponse(data, safe=False)


@staff_member_required
def get_customer_detail(request, customer_id):
    """Get single customer detail as JSON"""
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        orders = Order.objects.filter(customer_email=customer.email).order_by('-created_at')
        
        data = {
            'id': customer.id,
            'name': customer.name,
            'email': customer.email,
            'phone': customer.phone,
            'address': customer.address,
            'district': customer.district,
            'is_active': customer.is_active,
            'is_diabetic': customer.is_diabetic,
            'diabetes_type': customer.diabetes_type,
            'has_high_blood_pressure': customer.has_high_blood_pressure,
            'blood_pressure_status': customer.blood_pressure_status,
            'order_count': orders.count(),
            'total_spent': sum(float(o.total_amount) for o in orders),
            'created_at': customer.created_at.isoformat() if customer.created_at else None,
            'orders': [
                {
                    'order_id': o.order_id,
                    'total_amount': float(o.total_amount),
                    'status': o.status,
                    'status_display': o.get_status_display(),
                    'created_at': o.created_at.isoformat() if o.created_at else None,
                    'items_count': o.get_total_items(),
                }
                for o in orders[:10]
            ]
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
@staff_member_required
def create_customer(request):
    """Create a new customer"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        name = data.get('name', '').strip()
        email = data.get('email', '').strip()
        phone = data.get('phone', '').strip()
        address = data.get('address', '').strip()
        district = data.get('district', '').strip()
        is_active = data.get('is_active', True)
        is_diabetic = data.get('is_diabetic', False)
        diabetes_type = data.get('diabetes_type', 'none')
        has_high_blood_pressure = data.get('has_high_blood_pressure', False)
        blood_pressure_status = data.get('blood_pressure_status', 'normal')
        
        if not name:
            return JsonResponse({'status': 'error', 'message': 'Name is required'}, status=400)
        if not email:
            return JsonResponse({'status': 'error', 'message': 'Email is required'}, status=400)
        
        # Check if customer already exists
        if Customer.objects.filter(email=email).exists():
            return JsonResponse({'status': 'error', 'message': 'Customer with this email already exists'}, status=400)
        
        customer = Customer.objects.create(
            name=name,
            email=email,
            phone=phone,
            address=address,
            district=district,
            is_active=is_active,
            is_diabetic=is_diabetic,
            diabetes_type=diabetes_type,
            has_high_blood_pressure=has_high_blood_pressure,
            blood_pressure_status=blood_pressure_status,
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Customer created successfully',
            'customer_id': customer.id
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@staff_member_required
def update_customer(request, customer_id):
    """Update an existing customer"""
    if request.method not in ['PUT', 'POST']:
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        data = json.loads(request.body)
        
        customer.name = data.get('name', customer.name).strip()
        customer.email = data.get('email', customer.email).strip()
        customer.phone = data.get('phone', customer.phone).strip()
        customer.address = data.get('address', customer.address).strip()
        customer.district = data.get('district', customer.district).strip()
        customer.is_active = data.get('is_active', customer.is_active)
        customer.is_diabetic = data.get('is_diabetic', customer.is_diabetic)
        customer.diabetes_type = data.get('diabetes_type', customer.diabetes_type)
        customer.has_high_blood_pressure = data.get('has_high_blood_pressure', customer.has_high_blood_pressure)
        customer.blood_pressure_status = data.get('blood_pressure_status', customer.blood_pressure_status)
        customer.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Customer updated successfully',
            'customer_id': customer.id
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@staff_member_required
def delete_customer(request, customer_id):
    """Delete a customer"""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        customer.delete()
        return JsonResponse({'status': 'success', 'message': 'Customer deleted successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@staff_member_required
def get_customer_orders(request, customer_id):
    """Get customer orders"""
    try:
        customer = get_object_or_404(Customer, id=customer_id)
        orders = Order.objects.filter(customer_email=customer.email).order_by('-created_at')
        
        data = [{
            'order_id': o.order_id,
            'total_amount': float(o.total_amount),
            'status': o.status,
            'status_display': o.get_status_display(),
            'payment_status': o.payment_status,
            'created_at': o.created_at.isoformat() if o.created_at else None,
            'items_count': o.get_total_items(),
            'items': [
                {
                    'product_name': item.product_name,
                    'quantity': item.quantity,
                    'total_price': float(item.total_price),
                }
                for item in o.order_items.all()
            ]
        } for o in orders]
        
        return JsonResponse({
            'status': 'success',
            'customer_name': customer.name,
            'orders': data,
            'total_orders': orders.count(),
            'total_spent': sum(float(o.total_amount) for o in orders)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    

# ============================================
# CUSTOMER EXPORT VIEWS
# ============================================



@staff_member_required
def export_customers_excel(request):
    """
    Export all customers to Excel file
    """
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"

        # Define headers
        headers = [
            'ID', 'Name', 'Email', 'Phone', 'Address', 'District', 
            'Status', 'Diabetes', 'Diabetes Type', 'Blood Pressure',
            'BP Status', 'Order Count', 'Total Spent', 'Created At'
        ]

        # Style for header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Get customers with order data
        customers = Customer.objects.all().order_by('-created_at')
        
        # Write data
        for row_idx, customer in enumerate(customers, 2):
            # Get customer orders
            orders = Order.objects.filter(customer_email=customer.email)
            order_count = orders.count()
            total_spent = sum(float(o.total_amount) for o in orders)
            
            row_data = [
                customer.id,
                customer.name,
                customer.email,
                customer.phone or '',
                customer.address or '',
                customer.district or '',
                'Active' if customer.is_active else 'Inactive',
                'Yes' if customer.is_diabetic else 'No',
                customer.diabetes_type or 'none',
                'Yes' if customer.has_high_blood_pressure else 'No',
                customer.blood_pressure_status or 'normal',
                order_count,
                f"{total_spent:.2f}",
                customer.created_at.strftime('%Y-%m-%d %H:%M') if customer.created_at else ''
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = length
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=customers_export_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        wb.save(response)
        return response

    except Exception as e:
        return JsonResponse({
            'status': 'error', 
            'message': f'Failed to export customers to Excel: {str(e)}'
        }, status=500)


@staff_member_required
def export_customers_pdf(request):
    """
    Export all customers to PDF file
    """
    try:
        # Create buffer for PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        # Get customers with order data
        customers = Customer.objects.all().order_by('-created_at')
        
        # Calculate totals
        total_customers = customers.count()
        total_orders = 0
        total_revenue = 0
        
        customer_data = []
        for customer in customers:
            orders = Order.objects.filter(customer_email=customer.email)
            order_count = orders.count()
            total_spent = sum(float(o.total_amount) for o in orders)
            total_orders += order_count
            total_revenue += total_spent
            
            customer_data.append({
                'customer': customer,
                'order_count': order_count,
                'total_spent': total_spent
            })

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2C3E50'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_LEFT,
            leading=10
        )

        # Build story
        story = []
        
        # Title
        story.append(Paragraph("Customer Report", title_style))
        story.append(Spacer(1, 10))
        
        # Summary
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#34495E'),
            alignment=TA_LEFT,
            spaceAfter=8
        )
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", summary_style))
        story.append(Paragraph(f"Total Customers: {total_customers}", summary_style))
        story.append(Paragraph(f"Total Orders: {total_orders}", summary_style))
        story.append(Paragraph(f"Total Revenue: ৳{total_revenue:,.2f}", summary_style))
        story.append(Spacer(1, 15))

        # Prepare table data
        headers = [
            'ID', 'Name', 'Email', 'Phone', 'District', 
            'Status', 'Diabetes', 'BP', 'Orders', 'Total Spent', 'Created'
        ]
        
        table_data = [headers]
        
        for data in customer_data:
            customer = data['customer']
            row = [
                str(customer.id),
                customer.name[:25] if customer.name else '',
                customer.email[:25] if customer.email else '',
                customer.phone or '',
                customer.district or '',
                'Active' if customer.is_active else 'Inactive',
                'Diabetic' if customer.is_diabetic else 'Normal',
                'High BP' if customer.has_high_blood_pressure else 'Normal',
                str(data['order_count']),
                f"৳{data['total_spent']:.2f}",
                customer.created_at.strftime('%Y-%m-%d') if customer.created_at else ''
            ]
            table_data.append(row)

        # Create table
        table = Table(table_data, repeatRows=1)
        
        # Table style
        table_style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            
            # Borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            
            # Alternate row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ])
        
        # Column widths
        col_widths = [
            0.4*inch, 0.9*inch, 1.7*inch, 1.0*inch, 
            0.6*inch, 0.5*inch, 0.6*inch, 0.5*inch,
            0.5*inch, 0.8*inch, 0.7*inch
        ]
        table._argW = col_widths
        
        table.setStyle(table_style)
        story.append(table)
        
        # Footer
        story.append(Spacer(1, 15))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#7F8C8D'),
            alignment=TA_CENTER
        )
        story.append(Paragraph("Generated by SuperWater Admin Panel", footer_style))

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        # Create response
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=customers_export_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
        return response

    except Exception as e:
        return JsonResponse({
            'status': 'error', 
            'message': f'Failed to export customers to PDF: {str(e)}'
        }, status=500)

#------------------------Registration of USER--------------------#
# views.py - Add these registration views


# views.py - সম্পূর্ণ আপডেটেড কোড

from django.shortcuts import render, redirect
from django.contrib.auth import login as auth_login  # 🔥 এটা গুরুত্বপূর্ণ
from django.contrib.auth import authenticate
from django.contrib.auth import logout as auth_logout  # 🔥 logout-এর জন্যও
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from .models import User, UserProfile, Customer
import re
import traceback


# views.py - আপডেটেড register view

from django.shortcuts import render, redirect
from django.contrib.auth import login as auth_login
from django.contrib.auth import authenticate
from django.contrib.auth import logout as auth_logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_protect  # 🔥 এটা যোগ করুন
from .models import User, UserProfile, Customer
import re
import traceback


# views.py - সম্পূর্ণ আপডেটেড

from django.shortcuts import render, redirect
from django.contrib.auth import login as auth_login
from django.contrib.auth import authenticate
from django.contrib.auth import logout as auth_logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.http import HttpResponse
from .models import User, UserProfile, Customer
import re


@csrf_protect
@never_cache
@ensure_csrf_cookie
def register(request):
    """User registration view - Customer auto create"""
    print("=" * 60)
    print("🔵 REGISTER VIEW CALLED")
    print(f"Method: {request.method}")
    print("=" * 60)
    
    # যদি ইউজার ইতিমধ্যে লগইন করা থাকে
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        print("\n🟢 POST DATA RECEIVED:")
        
        # Get form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        district = request.POST.get('district', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        role = request.POST.get('role', 'user')
        terms = request.POST.get('terms', False)
        
        print(f"📧 Email: {email}")
        print(f"👤 First Name: {first_name}")
        print(f"👤 Last Name: {last_name}")
        print(f"📱 Phone: {phone}")
        print(f"📍 Address: {address}")
        print(f"🏙️ District: {district}")
        print(f"✅ Terms: {terms}")
        
        # Validation
        errors = []
        
        # Check if terms accepted
        if not terms:
            errors.append('You must agree to the Terms of Service')
            print("❌ Terms not accepted")
        
        # Validate first name
        if len(first_name) < 2:
            errors.append('First name must be at least 2 characters long')
            print("❌ First name too short")
        
        # Validate last name
        if len(last_name) < 2:
            errors.append('Last name must be at least 2 characters long')
            print("❌ Last name too short")
        
        # Validate email
        try:
            validate_email(email)
            print("✅ Email validation passed")
        except ValidationError:
            errors.append('Please enter a valid email address')
            print("❌ Email validation failed")
        
        # Check if email exists
        print(f"\n🔍 Checking if email exists: {email}")
        
        if User.objects.filter(email__iexact=email).exists():
            existing_user = User.objects.filter(email__iexact=email).first()
            print(f"❌ Email already exists: {email} (User ID: {existing_user.id})")
            errors.append('An account with this email already exists. Please login or use a different email.')
        else:
            print("✅ Email is available for registration")
        
        # Validate phone
        phone_pattern = r'^[0-9+\-\s()]{10,15}$'
        if not re.match(phone_pattern, phone):
            errors.append('Please enter a valid phone number (10-15 digits)')
            print("❌ Invalid phone number")
        else:
            print("✅ Phone validation passed")
        
        # Validate address
        if len(address) < 5:
            errors.append('Address must be at least 5 characters long')
            print("❌ Address too short")
        else:
            print("✅ Address validation passed")
        
        # Validate district
        if len(district) < 2:
            errors.append('Please enter a valid district name')
            print("❌ District too short")
        else:
            print("✅ District validation passed")
        
        # Validate password
        if len(password) < 8:
            errors.append('Password must be at least 8 characters long')
            print("❌ Password too short")
        elif password != confirm_password:
            errors.append('Passwords do not match')
            print("❌ Passwords don't match")
        else:
            print("✅ Password length OK")
            # Password strength checks
            if not re.search(r'[A-Z]', password):
                errors.append('Password must contain at least one uppercase letter')
                print("❌ No uppercase letter")
            if not re.search(r'[a-z]', password):
                errors.append('Password must contain at least one lowercase letter')
                print("❌ No lowercase letter")
            if not re.search(r'[0-9]', password):
                errors.append('Password must contain at least one number')
                print("❌ No number")
            if not re.search(r'[^A-Za-z0-9]', password):
                errors.append('Password must contain at least one special character')
                print("❌ No special character")
        
        print(f"\n📊 Total Errors: {len(errors)}")
        if errors:
            print("Errors:")
            for error in errors:
                print(f"  ❌ {error}")
        
        # If no errors, create user and customer
        if not errors:
            print("\n" + "=" * 60)
            print("✅ ALL VALIDATIONS PASSED - Creating User and Customer")
            print("=" * 60)
            
            try:
                # ===== STEP 1: Create User =====
                print("\n🔄 Step 1: Creating User...")
                
                # 🔥🔥🔥 এখানে username=None যোগ করুন 🔥🔥🔥
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    username=None,  # ✅ এটা যোগ করুন
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone,
                    address=address,
                    district=district,
                    user_type=role,
                    is_active=True,
                    is_verified=False
                )
                print(f"✅ User created: {user.email} (ID: {user.id})")
                
                # ===== STEP 2: Create User Profile =====
                print("\n🔄 Step 2: Creating UserProfile...")
                UserProfile.objects.create(user=user)
                print(f"✅ UserProfile created for: {user.email}")
                
                # ===== STEP 3: Create Customer =====
                print("\n🔄 Step 3: Creating Customer...")
                customer = Customer.objects.create(
                    user=user,
                    name=f"{first_name} {last_name}",
                    email=email,
                    phone=phone,
                    address=address,
                    district=district,
                    is_active=True,
                    is_diabetic=False,
                    diabetes_type='none',
                    has_high_blood_pressure=False,
                    blood_pressure_status='normal'
                )
                print(f"✅ Customer created: {customer.name} (ID: {customer.id})")
                
                # ===== STEP 4: লগইন করে হোমে পাঠান =====
                print("\n🔄 Step 4: Logging in user...")
                auth_login(request, user)
                print("✅ User logged in")
                
                messages.success(
                    request, 
                    '🎉 Account created successfully! Welcome to Superwater.'
                )
                
                print("\n" + "=" * 60)
                print("✅ REGISTRATION COMPLETE - Redirecting to home")
                print("=" * 60)
                
                return redirect('login')
                
            except IntegrityError as e:
                print(f"\n❌ IntegrityError: {str(e)}")
                import traceback
                print(traceback.format_exc())
                
                if 'UNIQUE constraint failed' in str(e):
                    messages.error(request, 'This email is already registered. Please login or use a different email.')
                else:
                    messages.error(request, f'Database error: {str(e)}')
                    
            except Exception as e:
                print(f"\n❌ Unexpected error: {str(e)}")
                import traceback
                print(traceback.format_exc())
                messages.error(request, 'An unexpected error occurred. Please try again.')
        else:
            # Display errors
            print("\n❌ Validation errors found - Showing messages to user")
            for error in errors:
                messages.error(request, error)
                print(f"  Showing error: {error}")
    
    # GET request - show registration form
    print("\n🔄 Rendering registration form")
    return render(request, 'registration.html')

def user_login(request):
    """User login view - Direct password check"""
    print("=" * 60)
    print("🔵 LOGIN VIEW CALLED")
    print(f"Method: {request.method}")
    print("=" * 60)
    
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '')
        
        print(f"📧 Login Email: {email}")
        print(f"🔑 Password: {password}")
        print(f"🔑 Password length: {len(password)}")
        
        if not email or not password:
            messages.error(request, 'Please enter both email and password')
            return render(request, 'registration/login.html')
        
        try:
            # সরাসরি ইউজার খুঁজুন
            user = User.objects.get(email=email)
            print(f"👤 User found: {user.email}")
            print(f"🔓 Is active: {user.is_active}")
            
            # সরাসরি পাসওয়ার্ড চেক করুন
            password_valid = user.check_password(password)
            print(f"🔐 Password valid: {password_valid}")
            
            if password_valid and user.is_active:
                # লগইন করুন
                auth_login(request, user)
                messages.success(request, f'Welcome back, {user.first_name}!')
                
                next_url = request.GET.get('next', 'home')
                print(f"✅ Login successful, redirecting to: {next_url}")
                return redirect(next_url)
            elif not password_valid:
                messages.error(request, 'Invalid password. Please try again.')
                print("❌ Password incorrect")
            else:
                messages.error(request, 'Your account is inactive.')
                print("❌ User is inactive")
                
        except User.DoesNotExist:
            print("❌ User not found")
            messages.error(request, 'No account found with this email address.')
        except Exception as e:
            print(f"❌ Error: {e}")
            print(traceback.format_exc())
            messages.error(request, 'An error occurred during login. Please try again.')
    
    return render(request, 'login.html')


def user_logout(request):
    """User logout view"""
    auth_logout(request)
    messages.info(request, 'You have been logged out successfully.')
    return redirect('home')


@login_required
def profile(request):
    """User profile view"""
    user = request.user
    
    if request.method == 'POST':
        # Update profile
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        district = request.POST.get('district', '').strip()
        
        errors = []
        
        if len(first_name) < 2:
            errors.append('First name must be at least 2 characters')
        if len(last_name) < 2:
            errors.append('Last name must be at least 2 characters')
        
        phone_pattern = r'^[0-9+\-\s()]{10,15}$'
        if not re.match(phone_pattern, phone):
            errors.append('Please enter a valid phone number')
        
        if len(address) < 5:
            errors.append('Address must be at least 5 characters')
        if len(district) < 2:
            errors.append('Please enter a valid district')
        
        if not errors:
            user.first_name = first_name
            user.last_name = last_name
            user.phone = phone
            user.address = address
            user.district = district
            user.save()
            
            # Update customer record as well
            try:
                customer = Customer.objects.get(email=user.email)
                customer.name = f"{first_name} {last_name}"
                customer.phone = phone
                customer.address = address
                customer.district = district
                customer.save()
            except Customer.DoesNotExist:
                pass
            
            messages.success(request, 'Profile updated successfully!')
            return redirect('profile')
        else:
            for error in errors:
                messages.error(request, error)
    
    return render(request, 'profile.html', {'user': user})


@login_required
def customer_dashboard(request):
    """Customer dashboard showing health profile and orders"""
    try:
        # Get or create customer record
        customer, created = Customer.objects.get_or_create(
            email=request.user.email,
            defaults={
                'user': request.user,
                'name': request.user.full_name,
                'phone': request.user.phone,
                'address': request.user.address,
                'district': request.user.district,
            }
        )
        # Update customer if user info changed
        if not created:
            customer.name = request.user.full_name
            customer.phone = request.user.phone
            customer.address = request.user.address
            customer.district = request.user.district
            customer.save()
        
        context = {
            'customer': customer,
            'health_summary': customer.get_health_summary() if hasattr(customer, 'get_health_summary') else [],
        }
        return render(request, 'registration/customer_dashboard.html', context)
    except Customer.DoesNotExist:
        # Create customer if doesn't exist
        customer = Customer.objects.create(
            user=request.user,
            name=request.user.full_name,
            email=request.user.email,
            phone=request.user.phone,
            address=request.user.address,
            district=request.user.district,
        )
        return redirect('customer_dashboard')


@login_required
def update_health_info(request):
    """Update customer health information"""
    if request.method == 'POST':
        try:
            customer = Customer.objects.get(email=request.user.email)
        except Customer.DoesNotExist:
            customer = Customer.objects.create(
                user=request.user,
                name=request.user.full_name,
                email=request.user.email,
            )
        
        # Update health info
        customer.is_diabetic = request.POST.get('is_diabetic') == 'on'
        customer.diabetes_type = request.POST.get('diabetes_type', 'none')
        customer.has_high_blood_pressure = request.POST.get('has_high_blood_pressure') == 'on'
        customer.blood_pressure_status = request.POST.get('blood_pressure_status', 'normal')
        customer.save()
        
        messages.success(request, 'Health information updated successfully!')
        return redirect('customer_dashboard')
    
    return redirect('customer_dashboard')


@login_required
def change_password(request):
    """Change password view"""
    if request.method == 'POST':
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        # Verify current password
        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect')
            return redirect('change_password')
        
        if len(new_password) < 8:
            messages.error(request, 'New password must be at least 8 characters long')
            return redirect('change_password')
        
        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match')
            return redirect('change_password')
        
        # Change password
        request.user.set_password(new_password)
        request.user.save()
        messages.success(request, 'Password changed successfully! Please login again.')
        
        # Logout and redirect to login
        auth_logout(request)
        return redirect('login')
    
    return render(request, 'change_password.html')


# views.py
from django.http import JsonResponse
from django.contrib.auth import get_user_model

User = get_user_model()

def check_auth(request):
    """Check if user is authenticated - Real time check"""
    print("=" * 50)
    print("🔍 CHECK AUTH CALLED")
    print(f"User: {request.user}")
    print(f"Is authenticated: {request.user.is_authenticated}")
    print(f"Session key: {request.session.session_key}")
    print("=" * 50)
    
    return JsonResponse({
        'is_authenticated': request.user.is_authenticated,
        'user_email': request.user.email if request.user.is_authenticated else None,
        'user_id': request.user.id if request.user.is_authenticated else None,
        'session_key': request.session.session_key
    })
#--------------------------------------Inventory add products--------------------#
# views.py


# ===== Add Inventory (Stock ইনক্রিজ) =====
# views.py
# views.py
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import Product, Inventory
import json
import csv
from datetime import datetime
from decimal import Decimal

# ===== Add Inventory API =====
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def add_inventory_api(request):
    """API: Add inventory by product name - Admin only"""
    try:
        # Check admin access
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        data = json.loads(request.body)
        product_name = data.get('product_name', '').strip()
        quantity = data.get('quantity', 0)
        notes = data.get('notes', '')
        
        if not product_name:
            return JsonResponse({'status': 'error', 'message': 'Product name is required'}, status=400)
        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': 'Quantity must be greater than 0'}, status=400)
        
        # Find product by name (case insensitive)
        product = Product.objects.filter(name__iexact=product_name, is_active=True).first()
        
        if not product:
            return JsonResponse({
                'status': 'error', 
                'message': f'Product "{product_name}" not found. Please check the name.'
            }, status=404)
        
        # Get current stock
        try:
            current_stock = product.get_inventory_total()
        except:
            current_stock = product.stock_quantity or 0
        
        new_stock = current_stock + quantity
        
        # Create inventory entry
        try:
            from .models import Inventory
            inventory = Inventory.objects.create(
                product=product,
                quantity=quantity,
                previous_stock=current_stock,
                new_stock=new_stock,
                movement_type='add',
                notes=notes or f'Added {quantity} units - Added by {request.user.email}',
                user=request.user
            )
        except:
            # Inventory model doesn't exist, update stock_quantity directly
            product.stock_quantity = new_stock
            product.save()
            inventory = None
        
        return JsonResponse({
            'status': 'success',
            'message': f'Added {quantity} units to {product.name}',
            'product_id': product.id,
            'product_name': product.name,
            'previous_stock': current_stock,
            'new_stock': new_stock
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)




# ===== Export Inventory Excel =====
@login_required
def export_inventory_excel(request):
    """Export inventory data as CSV/Excel"""
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Admin access required'}, status=403)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Product ID', 'Product Name', 'Category', 'Price', 
            'Stock Quantity', 'Inventory Total', 'Inventory Value', 
            'Status', 'Created At'
        ])
        
        products = Product.objects.filter(is_active=True).order_by('-created_at')
        
        for product in products:
            try:
                inventory_total = product.get_inventory_total()
            except:
                inventory_total = product.stock_quantity or 0
            
            inventory_value = float(product.price) * inventory_total
            
            status = 'In Stock' if inventory_total > 20 else 'Low Stock' if inventory_total > 0 else 'Out of Stock'
            
            writer.writerow([
                product.id,
                product.name,
                product.category,
                float(product.price),
                product.stock_quantity,
                inventory_total,
                inventory_value,
                status,
                product.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ===== Export Inventory PDF =====
@login_required
def export_inventory_pdf(request):
    """Export inventory data as PDF"""
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({'status': 'error', 'message': 'Admin access required'}, status=403)
        
        # Check if reportlab is available
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            return JsonResponse({
                'status': 'error', 
                'message': 'ReportLab not installed. Please install: pip install reportlab'
            }, status=400)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#007bff'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        story = []
        
        # Title
        story.append(Paragraph("📦 Inventory Report - Superwater", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", 
                              ParagraphStyle('SubTitle', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#666'), alignment=TA_CENTER)))
        story.append(Spacer(1, 0.2*inch))
        
        # Table data
        table_data = [['ID', 'Product', 'Category', 'Price', 'Stock', 'Inventory', 'Value', 'Status']]
        
        products = Product.objects.filter(is_active=True).order_by('-created_at')
        
        for product in products[:100]:  # Limit to 100 products
            try:
                inventory_total = product.get_inventory_total()
            except:
                inventory_total = product.stock_quantity or 0
            
            inventory_value = float(product.price) * inventory_total
            
            if inventory_total > 20:
                status = 'In Stock'
                status_color = 'green'
            elif inventory_total > 0:
                status = 'Low Stock'
                status_color = 'orange'
            else:
                status = 'Out of Stock'
                status_color = 'red'
            
            table_data.append([
                str(product.id),
                product.name[:25] + ('...' if len(product.name) > 25 else ''),
                product.category,
                f"৳{float(product.price):.2f}",
                str(product.stock_quantity),
                str(inventory_total),
                f"৳{inventory_value:.2f}",
                status
            ])
        
        # Create table
        table = Table(table_data, colWidths=[0.4*inch, 1.5*inch, 0.8*inch, 0.7*inch, 0.6*inch, 0.6*inch, 0.8*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ddd')),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
            ('ALIGN', (7, 1), (7, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 3),
        ]))
        
        story.append(table)
        
        # Footer
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("© 2026 Superwater - All rights reserved", 
                              ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#999'), alignment=TA_CENTER)))
        
        doc.build(story)
        buffer.seek(0)
        response.write(buffer.getvalue())
        return response
        
    except Exception as e:
        print(f"❌ PDF Export error: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
# ===== Remove Inventory (Stock ডিক্রিজ) =====
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def remove_inventory_api(request):
    """Remove stock from inventory - Admin only"""
    try:
        # Check admin access
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        data = json.loads(request.body)
        product_id = data.get('product_id')
        quantity = data.get('quantity', 0)
        notes = data.get('notes', '')
        
        if not product_id:
            return JsonResponse({'status': 'error', 'message': 'Product ID required'}, status=400)
        
        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': 'Quantity must be greater than 0'}, status=400)
        
        # Get product
        try:
            product = Product.objects.get(id=product_id, is_active=True)
        except Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)
        
        # Get current stock
        current_stock = product.get_inventory_total()
        
        if current_stock < quantity:
            return JsonResponse({
                'status': 'error', 
                'message': f'Insufficient stock. Available: {current_stock}, Requested: {quantity}'
            }, status=400)
        
        new_stock = current_stock - quantity
        
        # Create inventory entry
        inventory = Inventory.objects.create(
            product=product,
            quantity=-quantity,  # Negative for removal
            previous_stock=current_stock,
            new_stock=new_stock,
            movement_type='remove',
            notes=notes or f'Removed {quantity} units from inventory',
            user=request.user
        )
        
        return JsonResponse({
            'status': 'success',
            'message': f'Removed {quantity} units from {product.name}',
            'data': {
                'inventory_id': inventory.id,
                'product_id': product.id,
                'product_name': product.name,
                'previous_stock': current_stock,
                'removed_quantity': quantity,
                'new_stock': new_stock,
                'movement_type': 'remove'
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"Error: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ===== Get Inventory Status =====
@login_required
def get_inventory_status(request):
    """Get current inventory status - Admin only"""
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        products = Product.objects.filter(is_active=True)
        
        data = []
        for product in products:
            current_stock = product.get_inventory_total()
            data.append({
                'id': product.id,
                'name': product.name,
                'price': float(product.price),
                'current_stock': current_stock,
                'last_movement': product.get_inventory_movements().first().created_at.strftime('%Y-%m-%d %H:%M') if product.get_inventory_movements().exists() else None,
                'image_url': product.image.url if product.image else None
            })
        
        return JsonResponse({
            'status': 'success',
            'products': data,
            'total_products': len(data)
        })
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# ===== Get Inventory History =====
@login_required
def get_inventory_history(request, product_id):
    """Get inventory history for a specific product - Admin only"""
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        product = Product.objects.get(id=product_id, is_active=True)
        movements = product.get_inventory_movements()
        
        history = []
        for movement in movements:
            history.append({
                'id': movement.id,
                'quantity': movement.quantity,
                'previous_stock': movement.previous_stock,
                'new_stock': movement.new_stock,
                'movement_type': movement.movement_type,
                'notes': movement.notes,
                'created_at': movement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'user': movement.user.email if movement.user else 'System'
            })
        
        return JsonResponse({
            'status': 'success',
            'product_id': product.id,
            'product_name': product.name,
            'current_stock': product.get_inventory_total(),
            'history': history,
            'total_movements': len(history)
        })
        
    except Product.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import Product, Inventory
import json
from decimal import Decimal

# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import Product, Inventory
from decimal import Decimal
import json

# views.py
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def add_product_api(request):
    """API: Add inventory for existing product - Admin only"""
    try:
        # Check admin access
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        # Get form data
        product_id = request.POST.get('product_id')
        quantity = request.POST.get('quantity')
        notes = request.POST.get('notes', '')
        movement_type = request.POST.get('movement_type', 'add')
        
        # Validation
        if not product_id:
            return JsonResponse({'status': 'error', 'message': 'Product ID is required'}, status=400)
        if not quantity or int(quantity) <= 0:
            return JsonResponse({'status': 'error', 'message': 'Valid quantity is required'}, status=400)
        
        # Get existing product
        try:
            product = Product.objects.get(id=product_id, is_active=True)
        except Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Product not found'}, status=404)
        
        # Get current stock
        current_stock = product.get_inventory_total()
        new_stock = current_stock + int(quantity)
        
        # ✅ শুধু Inventory entry তৈরি করুন (Product তৈরি করবেন না)
        inventory = Inventory.objects.create(
            product=product,
            quantity=int(quantity),
            previous_stock=current_stock,
            new_stock=new_stock,
            movement_type=movement_type,
            notes=notes or f'Added {quantity} units to inventory - Added by {request.user.email}',
            user=request.user
        )
        print(f"✅ Inventory entry created: {inventory.id} for {product.name}")
        
        return JsonResponse({
            'status': 'success',
            'message': f'Added {quantity} units to {product.name} inventory',
            'inventory_id': inventory.id,
            'product_id': product.id,
            'product_name': product.name,
            'previous_stock': current_stock,
            'new_stock': new_stock
        })
        
    except Exception as e:
        print(f"❌ Error adding inventory: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)




# views.py - এই ফাংশনটি যোগ করুন
@csrf_exempt
@require_http_methods(["POST"])
@login_required
def add_product_by_name_api(request):
    """
    🔥 admin_inventory থেকে প্রোডাক্ট যোগ করলে:
    - শুধু ইনভেন্টরি বাড়বে
    - প্রোডাক্ট admin_products এ যোগ হবে না (is_active = False)
    - Admin পরে চাইলে admin_products এ active করতে পারবে
    """
    try:
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        data = json.loads(request.body)
        product_name = data.get('name', '').strip()
        quantity = data.get('quantity', 0)
        price = data.get('price', 0)
        description = data.get('description', '')
        
        if not product_name:
            return JsonResponse({'status': 'error', 'message': 'Product name is required'}, status=400)
        
        if quantity <= 0:
            return JsonResponse({'status': 'error', 'message': 'Quantity must be greater than 0'}, status=400)
        
        # চেক করুন প্রোডাক্ট ইতিমধ্যে আছে কিনা
        existing_product = Product.objects.filter(name__iexact=product_name).first()
        
        if existing_product:
            # প্রোডাক্ট থাকলে শুধু ইনভেন্টরি বাড়বে
            current_inventory = existing_product.get_inventory_total()
            new_inventory = current_inventory + quantity
            
            Inventory.objects.create(
                product=existing_product,
                quantity=quantity,
                previous_stock=current_inventory,
                new_stock=new_inventory,
                movement_type='add',
                notes=f'Added {quantity} units from admin_inventory - Added by {request.user.email}',
                user=request.user
            )
            
            return JsonResponse({
                'status': 'success',
                'message': f'Added {quantity} units to inventory: {product_name}',
                'product_id': existing_product.id,
                'product_name': existing_product.name,
                'inventory_total': new_inventory,
                'stock_quantity': existing_product.stock_quantity,
                'total_available': new_inventory + existing_product.stock_quantity,
                'action': 'inventory_added'
            })
        else:
            # 🔥 নতুন প্রোডাক্ট তৈরি করুন কিন্তু admin_products এ দেখাবেন না
            product = Product(
                name=product_name,
                price=Decimal(str(price)),
                stock_quantity=0,
                description=description,
                category='Filters',
            )
            # 🔥 is_active = False (Admin পরে active করতে পারবে)
            product.is_active = False
            product.save()
            
            print(f"✅ New product created in inventory: {product_name} (is_active=False)")
            
            # Inventory তৈরি করুন
            Inventory.objects.create(
                product=product,
                quantity=quantity,
                previous_stock=0,
                new_stock=quantity,
                movement_type='product_created',
                notes=f'Product created with {quantity} units from admin_inventory (inactive in admin_products)',
                user=request.user
            )
            
            return JsonResponse({
                'status': 'success',
                'message': f'Product "{product_name}" added to inventory with {quantity} units!',
                'product_id': product.id,
                'product_name': product.name,
                'inventory_total': quantity,
                'stock_quantity': 0,
                'total_available': quantity,
                'action': 'product_created_inventory_only',
                'note': 'This product is inactive in admin_products. Admin can activate it manually.'
            })
        
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# views.py - সম্পূর্ণ আপডেটেড ফাংশন

@csrf_exempt
@require_http_methods(["POST"])
@login_required
def add_product_from_admin_products(request):
    """
    🔥 admin_products থেকে প্রোডাক্ট যোগ করলে:
    - stock_quantity বাড়বে (Product Model)
    - inventory কমবে (Inventory Model)
    - প্রতিবার inventory চেক করবে
    """
    try:
        print("=" * 60)
        print("🔵 ADD PRODUCT FROM ADMIN PRODUCTS CALLED")
        
        # Check admin access
        if not request.user.is_staff and not request.user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Admin access required'
            }, status=403)
        
        # Get data from POST
        product_name = request.POST.get('name', '').strip()
        stock_quantity = request.POST.get('stock_quantity', 0)
        price = request.POST.get('price', 0)
        description = request.POST.get('description', '')
        product_id = request.POST.get('product_id')
        is_active = request.POST.get('is_active', 'true') == 'true'
        special_offer = request.POST.get('special_offer', '')
        discount_price = request.POST.get('discount_price', '')
        discount_percentage = request.POST.get('discount_percentage', '')
        
        print(f"📦 Product: {product_name}")
        print(f"📦 Stock Qty: {stock_quantity}")
        print(f"📦 Price: {price}")
        print(f"📦 Product ID: {product_id}")
        
        # Validation
        if not product_name:
            return JsonResponse({
                'status': 'error',
                'message': 'Product name is required'
            }, status=400)
        
        try:
            stock_quantity = int(stock_quantity)
            if stock_quantity <= 0:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Stock quantity must be greater than 0'
                }, status=400)
        except ValueError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid stock quantity'
            }, status=400)
        
        # ============================================================
        # 🔥 STEP 1: INVENTORY CHECK - প্রতিবার চেক করবে
        # ============================================================
        from django.db import models
        total_inventory = Inventory.objects.aggregate(
            total=models.Sum('quantity')
        )['total'] or 0
        
        print(f"📊 Total inventory available: {total_inventory}")
        print(f"📊 Requested stock: {stock_quantity}")
        
        # 🔥 Check if sufficient inventory
        if total_inventory < stock_quantity:
            print(f"❌ Insufficient inventory: {total_inventory} < {stock_quantity}")
            return JsonResponse({
                'status': 'error',
                'message': f'Insufficient inventory. Available: {total_inventory}, Required: {stock_quantity}'
            }, status=400)
        
        print(f"✅ Sufficient inventory available: {total_inventory} >= {stock_quantity}")
        
        # Find existing product
        existing_product = None
        if product_id:
            try:
                existing_product = Product.objects.get(id=product_id)
                print(f"✅ Found product by ID: {existing_product.name}")
            except Product.DoesNotExist:
                print(f"⚠️ Product ID {product_id} not found")
        
        if not existing_product:
            existing_product = Product.objects.filter(name__iexact=product_name).first()
            if existing_product:
                print(f"✅ Found product by name: {existing_product.name}")
            else:
                print(f"⚠️ No product found with name: {product_name}")
        
        if existing_product:
            # ============================================================
            # 🔥 UPDATE EXISTING PRODUCT
            # ============================================================
            print("🔄 Updating existing product...")
            
            # Update stock_quantity
            old_stock = existing_product.stock_quantity or 0
            new_stock = old_stock + stock_quantity
            existing_product.stock_quantity = new_stock
            
            if price:
                existing_product.price = Decimal(str(price))
            if description:
                existing_product.description = description
            if special_offer:
                existing_product.special_offer = special_offer
            if discount_price:
                existing_product.discount_price = Decimal(str(discount_price))
            if discount_percentage:
                existing_product.discount_percentage = int(discount_percentage)
            
            existing_product.is_active = is_active
            existing_product.save()
            print(f"✅ Stock updated: {old_stock} → {new_stock}")
            
            # ============================================================
            # 🔥 REDUCE INVENTORY (Inventory কমবে)
            # ============================================================
            current_inventory = existing_product.get_inventory_total()
            print(f"📊 Current inventory: {current_inventory}")
            
            if current_inventory >= stock_quantity:
                inventory = Inventory.objects.create(
                    product=existing_product,
                    quantity=-stock_quantity,
                    previous_stock=current_inventory,
                    new_stock=current_inventory - stock_quantity,
                    movement_type='remove',
                    notes=f'Removed {stock_quantity} from inventory for product: {product_name}',
                    user=request.user
                )
                remaining_inventory = current_inventory - stock_quantity
                removed_from_inventory = stock_quantity
                print(f"✅ Inventory reduced by {stock_quantity}")
                print(f"✅ Remaining inventory: {remaining_inventory}")
            elif current_inventory > 0:
                inventory = Inventory.objects.create(
                    product=existing_product,
                    quantity=-current_inventory,
                    previous_stock=current_inventory,
                    new_stock=0,
                    movement_type='remove',
                    notes=f'Removed all inventory ({current_inventory}) for product: {product_name}',
                    user=request.user
                )
                remaining_inventory = 0
                removed_from_inventory = current_inventory
                print(f"✅ All inventory removed: {current_inventory}")
            else:
                remaining_inventory = 0
                removed_from_inventory = 0
                print(f"⚠️ No inventory to remove")
            
            return JsonResponse({
                'status': 'success',
                'message': f'Added {stock_quantity} to stock: {product_name} (Inventory: -{removed_from_inventory})',
                'product_id': existing_product.id,
                'product_name': existing_product.name,
                'stock_quantity': new_stock,
                'inventory_total': remaining_inventory,
                'removed_from_inventory': removed_from_inventory,
                'action': 'stock_added_inventory_removed'
            })
        else:
            # ============================================================
            # 🔥 CREATE NEW PRODUCT
            # ============================================================
            print("🆕 Creating new product...")
            
            # Create product
            product = Product.objects.create(
                name=product_name,
                price=Decimal(str(price)),
                stock_quantity=stock_quantity,
                description=description,
                category='Filters',
                is_active=is_active,
                special_offer=special_offer if special_offer else None
            )
            print(f"✅ Product created: {product.name}")
            
            # Reduce inventory
            inventory = Inventory.objects.create(
                product=product,
                quantity=-stock_quantity,
                previous_stock=total_inventory,
                new_stock=total_inventory - stock_quantity,
                movement_type='remove',
                notes=f'Stock taken from inventory for new product: {product.name}',
                user=request.user
            )
            print(f"✅ Inventory reduced by {stock_quantity}")
            print(f"✅ Remaining inventory: {total_inventory - stock_quantity}")
            
            return JsonResponse({
                'status': 'success',
                'message': f'Product "{product_name}" created with {stock_quantity} units! (Inventory: -{stock_quantity})',
                'product_id': product.id,
                'product_name': product.name,
                'stock_quantity': stock_quantity,
                'inventory_total': total_inventory - stock_quantity,
                'removed_from_inventory': stock_quantity,
                'action': 'product_created_inventory_removed'
            })
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)